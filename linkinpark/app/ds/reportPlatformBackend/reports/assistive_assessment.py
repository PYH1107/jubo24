"""
ReportName: 輔具評估表
POC: Shen Chiang
"""

from datetime import datetime
from functools import reduce
from io import BytesIO

import pandas as pd
from bson.objectid import ObjectId
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from linkinpark.app.ds.reportPlatformBackend.utils import (
    ReportEmptyError, ReportGenerateError, get_nis_data, preprocess_date)


def assistive_assessment(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a statistic report of the assistive device
    recommended to use for patients in institution.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """
    # Regulation of this report
    if org != ObjectId("60ebc7a55738dd0028cd6bf9"):
        raise ReportGenerateError(
            "此為彰化養護專屬報表，因為院區設定問題、無法應用於其他機構。")

    # Querying section
    _ = suffix
    start, end, query_start, query_end = preprocess_date([start, end])
    query_setting = {
        "collection": [
            "assistivetechnologyassessments",
            "patients",
            "transfermanages",
            "organizations",
        ],
        "condition": [
            # 輔具評估
            {"organization": org,
             "createdDate": {"$gte": query_start, "$lt": query_end},
             "formMode": "assistiveTechnology"},
            # 住民檔案
            {"organization": org, "isDeleted": {"$ne": True}},
            # 住民異動
            {"organization": org, "createdDate": {"$lt": query_end}},
            # 機構檔案
            {"_id": org},
        ],
        "columns": [
            # 輔具評估
            {"_id": 0,
             "personalMobilityAids": 1,
             "orthoticsAndProsthetics": 1,
             "decompressionAids": 1,
             "foodAids": 1,
             "bathroomAccessories": 1,
             "dressingAids": 1,
             "communicationAndInformationAids": 1,
             "others": 1,
             "recommendations": 1,
             "patient": 1,
             "createdDate": 1},
            # 住民檔案
            {"lastName": 1, "firstName": 1},
            # 住民異動
            {"_id": 0,
             "patient": 1,
             "branch": 1,
             "room": 1,
             "bed": 1,
             "createdDate": 1},
            {"name": 1, "nickName": 1}
        ],
    }
    aid_df, patient_df, trans_df, org_df = get_nis_data(
        query_setting["collection"],
        query_setting["condition"],
        query_setting["columns"],
    )

    if aid_df.empty:
        raise ReportEmptyError("查詢區間內沒有相關輔具評估紀錄。")

    # Preprocess section
    aid_df = aid_df.where(pd.notnull(aid_df), None)
    exclude_col = ("recommendations", "patient", "createdDate")
    if "others" not in aid_df.columns:
        aid_df["others"] = None
    for col in aid_df.columns:
        if col in exclude_col:
            continue
        elif col in ("others",):
            aid_df[col] = aid_df[col].str.split(r"[,，、-]")
            aid_df.loc[aid_df[col].isnull(), "others"] = pd.Series(
                [[]] * len(aid_df[col]))
        aid_df[col + "_count"] = aid_df[col].apply(len)
        aid_df[col] = aid_df[col].apply(str)
        aid_df[col] = aid_df[col].str.replace(r"[\[\]']", "", regex=True)
    aid_df["evalDate"] = aid_df["createdDate"].dt.date
    patient_df.rename(columns={"_id": "patient"}, inplace=True)
    patient_df["name"] = patient_df["lastName"] + patient_df["firstName"]
    trans_df.sort_values(["patient", "createdDate"], inplace=True)
    trans_df = trans_df.loc[
        trans_df.groupby("patient")["createdDate"].idxmax()
    ]
    trans_df["room_bed"] = trans_df["room"] + trans_df["bed"]
    trans_df["branch_g"] = trans_df["branch"].copy()
    trans_df.loc[
        trans_df["branch_g"].isin(["1C", "1D", "1E"]), "branch_g"
    ] = "康樂家園"
    trans_df['branch_g'] = trans_df['branch_g'].astype("category")
    trans_df['branch_g'] = trans_df['branch_g'].cat.reorder_categories(
        ["康樂家園", "2C", "2D", "2E", "3C", "3D", "3E"]
    )

    # Merging section
    result = reduce(
        lambda l_df, r_df: pd.merge(l_df, r_df, on='patient', how='left'),
        [aid_df, patient_df, trans_df]
    )

    # Computing section
    detail_col = {
        "branch": "院區別",
        "room_bed": "房床號",
        "name": "姓名",
        "evalDate": "評估日期",
        "personalMobilityAids": "個人行動輔具",
        "orthoticsAndProsthetics": "矯具與義具",
        "decompressionAids": "減壓輔具",
        "foodAids": "飲食用輔具",
        "bathroomAccessories": "衛浴類輔具",
        "dressingAids": "衣著穿脫輔具",
        "communicationAndInformationAids": "溝通輔具",
        "others": "其他",
        "recommendations": "建議事項",
    }
    detail_df = result[detail_col].rename(columns=detail_col)
    stat_df = result.groupby("branch_g").sum().reset_index()
    stat_df.rename(columns={
        "branch_g": "區別",
        "personalMobilityAids_count": "個人行動輔具",
        "orthoticsAndProsthetics_count": "矯具與義具",
        "decompressionAids_count": "減壓輔具",
        "foodAids_count": "飲食用輔具",
        "bathroomAccessories_count": "衛浴類輔具",
        "dressingAids_count": "衣著穿脫輔具",
        "communicationAndInformationAids_count": "溝通輔具",
        "others_count": "其他",
    }, inplace=True)

    # Filing section
    report_file = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "輔具統計表"

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        stat_df.to_excel(
            writer,
            sheet_name="輔具統計表",
            index=False,
            header=True,
            startrow=2,
            startcol=0)
        detail_df.to_excel(
            writer,
            sheet_name="輔具明細表",
            index=False,
            header=True,
            startrow=2,
            startcol=0)
    workbook.close()

    # Formatting section
    wb = load_workbook(report_file)
    ws1 = wb["輔具統計表"]
    ws1["A1"] = (
        f"{org_df.at[0, 'name']}輔具評估統計表{start.year - 1911}年{start.month}月"
    )
    ws1["A1"].font = Font(size=16, bold=True)
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.merge_cells("A1:I1")
    for i in range(1, 10):
        ws1.column_dimensions[get_column_letter(i)].width = 14

    ws2 = wb["輔具明細表"]
    ws2["A1"] = (
        f"{org_df.at[0, 'name']}輔具評估明細表{start.year - 1911}年{start.month}月"
    )
    ws2["A1"].font = Font(size=16, bold=True)
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.merge_cells("A1:M1")
    for i in range(1, 13):
        ws2.column_dimensions[get_column_letter(i)].width = 14
    ws2.column_dimensions["M"].width = 14 * 4

    report_file = BytesIO()
    wb.save(report_file)
    wb.close()

    return report_file
