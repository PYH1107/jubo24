"""
ReportName: 疼痛監控月報
POC: Shen Chiang
"""

from datetime import datetime
from io import BytesIO

import pandas as pd
from bson.objectid import ObjectId
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from linkinpark.app.ds.reportPlatformBackend.utils import (
    clients_infile,
    ReportGenerateError,
    get_nis_data,
    preprocess_date,
    trans_timezone)


def pain_group(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate statistic of patient with pain score in
    different level groups and the case .
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """
    # Regulation of this report
    if org != ObjectId("60ebc7a55738dd0028cd6bf9"):
        raise ReportGenerateError(
            "此為彰化養護專屬報表，因為院區設定問題、無法應用於其他機構。")

    # Querying section
    _ = suffix
    start, end, query_start, query_end = preprocess_date([start, end])
    infile = clients_infile(start, end, org)["patient"].unique().tolist()
    trans_df, patient_df, vital_sign_df, case_df, org_df = get_nis_data(
        ["transfermanages",
         "patients",
         "vitalsigns",
         "eventanalyses",
         "organizations"],
        [{"organization": org,
          "createdDate": {"$lt": query_end},
          "status": {"$nin": ["reservation", "unpresented"]},
          "patient": {"$in": infile}},
         {"organization": org,
          "isDeleted": {"$ne": True},
          "_id": {"$in": infile}},
         {"organization": org,
          "createdDate": {"$gte": query_start,
                          "$lt": query_end},
          "patient": {"$in": infile}},
         {"organization": org,
          "resourceFormType": "Pain",
          "trackDate": {"$gte": query_start,
                        "$lt": query_end}},
         {"_id": org}],
        [{"_id": False,
          "patient": True,
          "branch": True,
          "room": True,
          "bed": True,
          "createdDate": True,
          "status": True},
         {"lastName": True,
          "firstName": True,
          "medicalRecordNote": True},
         {"patient": True,
          "createdDate": True,
          "PAIN": True},
         {"patient": True,
          "trackDate": True,
          "status": True,
          "reason": True,
          "improvementPlan": True},
         {"name": True}]
    )

    # Preprocess Section
    trans_df.sort_values(["patient", "createdDate"], inplace=True)
    trans_df["pre_status"] = trans_df.groupby("patient")["status"].shift(1)
    trans_df["pre_status_date"] = trans_df.groupby("patient")[
        "createdDate"].shift(1)
    trans_df = trans_df.loc[
        trans_df.groupby("patient")["createdDate"].idxmax()
    ]
    trans_df["branch_group"] = trans_df["branch"].copy()
    trans_df.loc[
        trans_df["branch_group"].isin(["1C", "1D", "1E"]), "branch_group"
    ] = "康樂家園"
    trans_df['branch_group'] = trans_df['branch_group'].astype("category")
    trans_df['branch_group'] = trans_df['branch_group'].cat.reorder_categories(
        ["康樂家園", "2C", "2D", "2E", "3C", "3D", "3E"]
    )
    patient_df.rename(columns={"_id": "patient"}, inplace=True)
    patient_df["medicalRecordNote"].fillna("", inplace=True)
    patient_df["palliativeCare"] = False
    patient_df.loc[
        patient_df["medicalRecordNote"].str.contains("安寧"), "palliativeCare"
    ] = True
    if vital_sign_df.empty:
        vital_sign_df = pd.DataFrame(columns=[
            "patient",
            "createdDate",
            "PAIN",
            "group",
        ])
        vital_sign_df["group"] = vital_sign_df["group"].astype("category")
        vital_sign_df["group"] = vital_sign_df["group"].cat.add_categories(
            ["疼痛指數≦3分", "疼痛指數≧4分"]
        )
    else:
        vital_sign_df = vital_sign_df.groupby(
            "patient").agg(score=("PAIN", "max")).reset_index()
        vital_sign_df["group"] = pd.cut(
            vital_sign_df["score"],
            bins=[-float("inf"), 3, float("inf")],
            labels=["疼痛指數≦3分", "疼痛指數≧4分"]
        )
    if case_df.empty:
        case_df = pd.DataFrame(columns=[
            "patient",
            "trackDate",
            "status",
            "reason",
            "improvementPlan",
        ])
    case_df["trackDate"] = trans_timezone(case_df["trackDate"], 0, 8)
    case_df.rename(columns={"status": "description"}, inplace=True)

    # Merging section
    upper_result = pd.merge(trans_df, vital_sign_df, "left", on="patient")
    upper_result = pd.merge(upper_result, patient_df, on="patient")
    lower_result = pd.merge(case_df, trans_df, "left", on="patient")
    lower_result = pd.merge(lower_result, patient_df, on="patient")

    # Computing section

    # 2022-07-20
    # If patient is in hospital for the whole report period, than the care
    # giver could not measure the patient's pain score. By institution's
    # definition, this kind of patient will be viewed as without pain.
    upper_result.loc[
        (upper_result["status"].isin(["unplannedHosp", "hospTransfer"]))
        & (upper_result["createdDate"] < query_start),
        "group"
    ] = "疼痛指數≦3分"
    upper_result.loc[
        (upper_result["status"] == "discharge")
        & (upper_result["pre_status"].isin(["unplannedHosp", "hospTransfer"]))
        & (upper_result["pre_status_date"] < query_start),
        "group"
    ] = "疼痛指數≦3分"

    upper_result["group"] = upper_result["group"].cat.add_categories(
        "尚未評估")
    upper_result['group'] = upper_result['group'].cat.reorder_categories([
        "疼痛指數≦3分", "疼痛指數≧4分", "尚未評估"
    ])
    upper_result.loc[pd.isna(upper_result["group"]), "group"] = "尚未評估"
    no_result = upper_result[upper_result["group"] == "尚未評估"]
    if len(no_result) > 0:
        patient_with_no_assessment = True
        col_names = {
            "branch": "區別",
            "room": "房號",
            "bed": "床號",
            "lastName": "姓氏",
            "firstName": "名字",
        }
        no_result = no_result[col_names.keys()]
        no_result.rename(columns=col_names, inplace=True)
    else:
        patient_with_no_assessment = False
    upper_result = upper_result.groupby(
        ["branch_group", "palliativeCare", "group"]
    ).size().reset_index()
    upper_result = upper_result.pivot(
        index=["branch_group", "group"], columns="palliativeCare", values=0
    ).reset_index()
    upper_result = upper_result.pivot(
        index="branch_group", columns="group"
    ).reset_index()
    upper_result.insert(1, "patients", upper_result.iloc[:, 1:].sum(axis=1))
    upper_result.columns = upper_result.columns.to_flat_index()
    upper_result.rename(
        columns={
            ('branch_group', ''): "區別",
            ('patients', ''): "住民總數",
            (False, '疼痛指數≦3分'): "疼痛指數≦3分（人數）",
            (False, '疼痛指數≧4分'): "疼痛指數≧4分（人數）",
            (False, '尚未評估'): "尚未評估",
            (True, '疼痛指數≦3分'): "疼痛指數≦3分（人數）",
            (True, '疼痛指數≧4分'): "疼痛指數≧4分（人數）",
            (True, '尚未評估'): "尚未評估",
        },
        inplace=True)
    lower_result["display_name"] = (
        lower_result["branch"]
        + " "
        + lower_result["room"]
        + "-"
        + lower_result["bed"]
        + lower_result["lastName"]
        + lower_result["firstName"])
    lower_result = lower_result[[
        "display_name", "trackDate", "description", "reason", "improvementPlan"
    ]]
    for col in range(3, 8, 2):
        lower_result.insert(col, "space" + str(col), None)
    lower_result.rename(columns={
        "display_name": "個案",
        "trackDate": "逐案分析日期",
        "description": "狀況說明",
        "reason": "原因分析",
        "improvementPlan": "改善措施",
    }, inplace=True)

    # Filing section
    report_file = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "疼痛監測月報"

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        upper_result.to_excel(
            writer,
            sheet_name="疼痛監測月報",
            index=False,
            header=True,
            startrow=3,
            startcol=0)
        lower_result.to_excel(
            writer,
            sheet_name="疼痛監測月報",
            index=False,
            header=True,
            startrow=12,
            startcol=0)
        if patient_with_no_assessment:
            no_result.to_excel(
                writer,
                sheet_name="尚未評估者名單",
                index=False,
                header=True,
            )
    workbook.close()

    # Formatting section
    # Close file written by pandas and reopen it with openpyxl to avoid unknown
    # crushes cause by end file problems while opening the file in Microsoft
    # Excel.
    workbook = load_workbook(report_file)
    worksheet = workbook["疼痛監測月報"]
    for col in range(1, 11):
        worksheet.column_dimensions[get_column_letter(col)].width = 15
    worksheet["A1"] = (
        f"{org_df.at[0, 'name']}疼痛監測月報表{start.year - 1911}年{start.month}月"
    )
    worksheet["A1"].font = Font(size=16, bold=True)
    worksheet["C3"] = "非安寧個案"
    worksheet["F3"] = "安寧個案"

    additional_cell_to_wrap = []
    additional_cell_to_merge = []
    if not lower_result.empty:
        for row in range(lower_result.shape[0]):
            for col in range(3, 8, 2):
                additional_cell_to_wrap.append(
                    get_column_letter(col) + str(row + 14)
                )
                additional_cell_to_merge.append(
                    get_column_letter(col)
                    + str(row + 14)
                    + ":"
                    + get_column_letter(col + 1)
                    + str(row + 14)
                )

    cell_to_bold = ("C3", "F3")
    for cell in cell_to_bold:
        worksheet[cell].font = Font(size=14, bold=True)

    cell_to_center = ("A1", "C3", "F3")
    for cell in cell_to_center:
        worksheet[cell].alignment = Alignment(horizontal="center")

    cell_to_wrap = ["C4", "D4", "F4", "G4"]
    cell_to_wrap += additional_cell_to_wrap
    for cell in cell_to_wrap:
        worksheet[cell].alignment = Alignment(wrapText=True)

    cell_to_merge = [
        "A1:H1",
        "C3:E3",
        "F3:H3",
        "C13:D13",
        "E13:F13",
        "G13:H13",
    ]
    cell_to_merge += additional_cell_to_merge
    for cell in cell_to_merge:
        worksheet.merge_cells(cell)
    report_file = BytesIO()
    workbook.save(report_file)
    workbook.close()

    return report_file
