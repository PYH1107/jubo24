"""
ReportName: 疼痛評估
POC: Shen Chiang
"""

from datetime import datetime
from io import BytesIO

import pandas as pd
from bson.objectid import ObjectId
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.page import PageMargins
from linkinpark.app.ds.reportPlatformBackend.utils import schema as dic
from linkinpark.app.ds.reportPlatformBackend.utils import (
    ReportEmptyError, get_nis_data, preprocess_date, trans_timezone)


def pain_level(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a list of clients who had pain suffering in
    the report period.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """
    # Querying section
    _ = suffix

    start, end, query_start, query_end = preprocess_date([start, end])

    patients = get_nis_data(
        "patients",
        {"organization": org, "isDeleted": {"$ne": True}},
        {"organization": 1,
         "lastName": 1,
         "firstName": 1,
         "room": 1,
         "bed": 1},
    ).rename(columns={"_id": "patient"})

    patients["pt_name"] = (
        patients["lastName"].fillna("") + patients["firstName"].fillna(""))
    patients["room_bed"] = (
        patients["room"].fillna("") + "-" + patients["bed"].fillna(""))

    org_name = get_nis_data(
        "organizations",
        {"_id": org},
        {"name": 1,
         "nickName": 1},
    ).rename(columns={"_id": "organization"})

    pains_col = {
        "part": 1,
        "partDetail": 1,
        "FRS": 1,
        "process": 1,
        "result": 1,
        "patient": 1,
        "createdDate": 1,
        "user": 1,
    }
    pains = get_nis_data(
        "pains",
        {"organization": org,
         "createdDate": {"$gte": query_start,
                         "$lt": query_end}},
        pains_col
    )
    if pains.empty:
        raise ReportEmptyError("查詢區間內查無相關疼痛評估紀錄")

    # Mechanism for enhance fault tolerance, which may deal with the situation
    # while there is no pain record found in the querying period.
    for col in pains_col.keys():
        if col not in pains.columns:
            pains[col] = None

    users = get_nis_data(
        "users",
        {"organization": org},
        {"lastName": 1,
         "firstName": 1}
    ).rename(columns={"_id": "user"})
    users["user_name"] = users["lastName"] + users["firstName"]

    # Merging section
    result = pd.merge(
        patients, org_name, how="left", on="organization"
    )
    result = pd.merge(
        pains, result, how="left", on="patient",
        # Only show those patient with pain evaluation records.
    )
    result = pd.merge(
        result, users, how="left", on="user",
    )

    # Formatting section
    result["createdDate"] = trans_timezone(
        result["createdDate"], from_utc=0, to_utc=8, ignore_nan=True,
    )
    result["date"] = result["createdDate"].dt.date
    result["time"] = result["createdDate"].dt.strftime("%H:%M")

    keep_column = [
        "pt_name",
        "room_bed",
        "date",
        "partDetail",
        "time",
        "FRS",
        "process",
        "result",
        "user_name"
    ]

    result = result[keep_column]

    replace_dict = {
        "result": dic.pain_result,
        "process": dic.pain_process,
    }
    result = result.replace(replace_dict)
    result["process"] = result["process"].fillna("")
    result["process"] = result["process"].apply(
        lambda x: [
            replace_dict["process"][item] if item in replace_dict["process"]
            else item for item in x])

    result["process"] = result["process"].apply(str)
    for pattern in ["[", "'", "]"]:
        result["process"] = result["process"].str.replace(
            pattern, "", regex=False)

    result = result.rename(
        columns={
            "pt_name": "住民",
            "room_bed": "床號",
            "date": "日期",
            "partDetail": "部位",
            "time": "發生時間",
            "FRS": "分數",
            "process": "護理措施",
            "result": "評值",
            "user_name": "評估者"
        }
    )

    # Filing section
    report_file = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "疼痛評估報表"

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        result.to_excel(writer,
                        sheet_name="疼痛評估報表",
                        index=False,
                        header=True,
                        startrow=4,
                        startcol=0)
    worksheet.page_margins = PageMargins(left=0.25,
                                         right=0.25,
                                         top=0.5,
                                         bottom=0.5,
                                         header=0.25,
                                         footer=0.25, )

    worksheet["A1"] = org_name["name"][0]
    worksheet["A3"] = "表單"
    worksheet["C3"] = "疼痛評估報表"
    worksheet["F3"] = "日期"
    worksheet["G3"] = (
        start.strftime("%Y-%m-%d")
        + " - "
        + (end - relativedelta(days=1)).strftime("%Y-%m-%d")
    )
    worksheet["A1"].font = Font(size=14, bold=True)
    worksheet["A3"].font = Font(size=12, bold=True)
    worksheet["F3"].font = Font(size=12, bold=True)
    worksheet["A1"].alignment = Alignment(
        horizontal="center", vertical="center"
    )

    n = len(result.index)
    for row in range(6, 6 + n):
        worksheet.cell(row=row, column=7).alignment = Alignment(wrapText=True)

    for column in range(1, 10):
        for row in range(6, 6 + n):
            if column == 7:
                continue
            worksheet.cell(row=row, column=column).alignment = Alignment(
                horizontal="center", vertical="top")

    for cell in ("A1:I2", "A3:B3", "C3:E3", "G3:I3", "A4:I4"):
        worksheet.merge_cells(cell)

    for column in ["B", "F"]:
        worksheet.column_dimensions[column].width = 6
    for column in ["C", "D", "E", "H"]:
        worksheet.column_dimensions[column].width = 10
    worksheet.column_dimensions["G"].width = 25

    report_file = BytesIO()
    workbook.save(report_file)
    workbook.close()

    return report_file
