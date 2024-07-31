"""
ReportName: 花蓮居服長期照顧差異報表
POC: Shen Chiang
"""
from datetime import datetime
from functools import reduce
from io import BytesIO
from os import path

import pandas as pd
from bson.objectid import ObjectId
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from linkinpark.app.ds.reportPlatformBackend.utils import (ReportGenerateError,
                                                           check_org_type,
                                                           count_age,
                                                           get_nis_data,
                                                           preprocess_date)


def hualien_homecare_community_different_report(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function is used to generate a report of patient with different
    status in the Hualien homecare community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    if not check_org_type(org, "homecare"):
        raise ReportGenerateError("此為居服專屬報表，無法應用於其他機構類型。")

    # parameter setting section
    _ = suffix
    start, end, query_start, query_end = preprocess_date([start, end])
    collection = [
        "patients",
        "servicemanagements",
        "daycareservices",
        "approvedcareplans",
        "approvedcareplans",
        "transfermanages",
    ]
    condition = [
        {"organization": org,
         "idDeleted": {"$ne": True},
         "DCHCServiceType": "homeCare"},
        {"organization": org,
         "start": {"$gte": query_start, "$lt": query_end},
         "funding": "subsidy"},
        {"organization": org},
        {"organization": org,
         "dateOfApproval": {'$gte': query_start},
         "planType": {"$ne": "termination"}},
        {"organization": org,
         "dateOfApproval": {'$lt': query_start},
         "planType": {"$ne": "termination"}},
        {"organization": org,
         "createdDate": {'$lt': query_end}},
    ]
    patient_col = {
        "birthday": 1,
        "numbering": 1,
        "caseNumber": 1,
        "sex": 1,
        "residentialAddress": 1,
        "DCHCServiceType": 1,
        "lastName": 1,
        "firstName": 1
    }
    care_plan_col = {
        "_id": 0,
        "CMSLevel": 1,
        "dateOfApproval": 1,
        "socialWelfareStatus": 1,
        "socialWelfare": 1,
        "patient": 1,
        "disability": 1
    }
    service_col = {
        "patient": 1, "start": 1, "service": 1
    }
    columns = [
        patient_col,
        service_col,
        {"code": 1},
        care_plan_col,
        care_plan_col,
        {"patient": 1, "createdDate": 1, "status": 1},
    ]

    replace_symbol = {"[": "", "]": "", "'": ""}
    cms_coordinate = dict((y, 5 + y * 3) for y in range(2, 9))
    sex_coordinate = {"男": 1, "女": 2}
    case_coordinate = {
        "65歲以上老人（含IADLs失能且獨居之老人）": 7,
        "65歲以上領身心障礙證明者": 11,
        "64歲以下領身心障礙證明者": 15,
        "55-64歲原住民": 19,
        "50歲以上失智症": 23,
        "無法判斷": 27,
    }
    social_coordinate = {"長照低收": 1, "長照中低收": 2, "一般戶": 3, }

    # Querying section
    (
        patient, service, code, careplan_now, careplan_pre, status
    ) = get_nis_data(collection, condition, columns)
    if careplan_now.empty:
        raise ReportGenerateError("查詢區間內無更新之服務計畫，故無法比較差異。")

    # Preprocess Section
    sheet_title = (
        f"中華民國{start.year - 1911}年{'上' if start.month <= 6 else '下'}半年"
        f"（{start.month}月至{end.month - 1}月）"
    )
    for col in patient_col:
        if col not in patient.columns:
            patient[col] = None
    if service.empty:
        service = pd.DataFrame(columns=service_col.keys())
    patient["name"] = patient["lastName"] + patient["firstName"]
    patient["age"] = count_age(patient["birthday"], datetime(start.year, 1, 1))
    patient["b_day"] = patient["birthday"].dt.strftime("%Y-%m-%d")
    address = patient["residentialAddress"].str.split(
        pat="([縣市鄉鎮區村里])", n=2, expand=True
    )
    patient["county"] = address[0] + address[1]
    patient["region"] = address[2] + address[3]
    patient["displayName"] = (
        "[" + patient["region"] + "] "
        + patient["numbering"].fillna("") + " "
        + patient["lastName"].fillna("")
        + patient["firstName"].fillna(""))
    patient.rename(columns={"_id": "patient"}, inplace=True)

    start_df = status.loc[
        status["status"] == "startServer",
        ["patient", "createdDate"]
    ].rename(columns={"createdDate": "start_at"})
    start_df = start_df.loc[
        start_df.groupby("patient")["start_at"].idxmax()]
    close_df = status.loc[
        status["status"] == "closed",
        ["patient", "createdDate"]
    ].rename(columns={"createdDate": "close_at"})
    close_df = close_df.loc[
        close_df.groupby("patient")["close_at"].idxmax()]
    last_df = status.loc[status.groupby("patient")["createdDate"].idxmax()]
    last_df = last_df[["patient", "status"]].rename(
        columns={"status": "last_status"})
    status = reduce(
        lambda left, right: pd.merge(
            left, right, on=['patient'], how='outer'),
        [start_df, close_df, last_df]
    )
    for col in ("start_at", "close_at"):
        status[col] = status[col].dt.date
    careplan_now = careplan_now.loc[
        careplan_now.sort_values("dateOfApproval").groupby("patient")[
            "dateOfApproval"].idxmax()
    ].reset_index(drop=True)
    careplan_pre = careplan_pre.loc[
        careplan_pre.sort_values("dateOfApproval").groupby("patient")[
            "dateOfApproval"].idxmax()
    ].reset_index(drop=True)
    code["category"] = code["code"].str.slice(0, 2)
    code.rename(columns={"_id": "service"}, inplace=True)

    # Merging section
    careplan = pd.merge(
        careplan_pre,
        careplan_now,
        "right",
        on="patient",
        suffixes=("_pre", "_now")
    )
    patient = pd.merge(patient, careplan, "left", on="patient")
    patient = pd.merge(patient, status, "left", on="patient")
    service = pd.merge(service, code, "left", on="service")

    # Computing section
    patient = patient[
        (patient["start_at"] >= start)
        | (~pd.isna(patient["dateOfApproval_now"]))]
    patient["reason"] = "照顧計畫更新"
    patient.loc[patient["start_at"] >= start, "reason"] = "開始服務"
    if patient.empty:
        raise ReportGenerateError("無法產製報表，因為區間內無服務計畫異動。")
    for row in patient.index:
        for stat in ("_pre", "_now"):
            col = [
                "reason",
                "disability" + stat,
                "age",
                "socialWelfareStatus" + stat
            ]
            reason, dementia, age, status = patient.loc[row, col]
            if isinstance(status, float):
                status = []
            write_to = "caseType" + stat
            if reason == "開始服務" and stat == "_pre":
                patient.loc[row, write_to] = None
            elif 50 <= age and dementia == "yes":
                patient.loc[row, write_to] = "50歲以上失智症"
            elif 55 <= age <= 64 and "indigenous" in status:
                patient.loc[row, write_to] = "55-64歲原住民"
            elif age <= 64 and "obstacle" in status:
                patient.loc[row, write_to] = "64歲以下領身心障礙證明者"
            elif 65 <= age and "obstacle" in status:
                patient.loc[row, write_to] = "65歲以上領身心障礙證明者"
            elif 65 <= age:
                patient.loc[row, write_to] = "65歲以上老人（含IADLs失能且獨居之老人）"
            else:
                patient.loc[row, write_to] = "無法判斷"
    for col in (
            "caseType", "CMSLevel", "disability", "socialWelfare"
    ):
        patient.loc[
            patient[col + "_pre"] != patient[col + "_now"], "diff"
        ] = True
    patient = patient[patient["diff"].fillna(False)]
    service_dict = {}
    for _id in service["patient"].unique():
        service_dict[_id] = {}
        for cat in ("BA", "GA"):
            service_date = list(
                service.loc[
                    (service["patient"] == _id)
                    & (service["category"] == cat),
                    "start"].dt.strftime("%Y-%m-%d").unique())
            service_dict[_id][cat] = sorted(service_date)
    service = pd.DataFrame.from_dict(
        service_dict, orient="index"
    ).reset_index().rename(columns={"index": "patient"})
    patient = pd.merge(patient, service, "left", on="patient")
    replace_dict = {
        "DCHCServiceType": {
            **replace_symbol,
            "HCOwnExpense": "自費",
            "HCRespite": "居家喘息",
            "homeCare": "居家服務",
        },
        **dict.fromkeys(
            ["socialWelfareStatus_pre", "socialWelfareStatus_now"],
            {**replace_symbol,
             "general": "一般戶",
             "SocialAssistanceLegalLowIncome": "社會救助法定低收入戶",
             "indigenous": "原住民",
             "lowIncome": "長照低收",
             "middleIncome": "長照中低收",
             "obstacle": "領有身心障礙證明",
             "veteran": "榮民"}
        ),
        **dict.fromkeys(
            ["socialWelfare_pre", "socialWelfare_now"],
            {**replace_symbol,
             "lowInLaw": "長照低收",
             "lowToMid": "長照中低收",
             "low": "長照低收",
             "normal": "一般戶",
             "nan": ""}
        ),
        "BA": {**replace_symbol, "nan": ""},
        "GA": {**replace_symbol, "nan": ""},
    }
    for col, replace in replace_dict.items():
        patient[col] = patient[col].astype(str)
        for k, v in replace.items():
            patient[col] = patient[col].str.replace(k, v, regex=True)
    patient.replace(
        {
            "sex": {"male": "男", "female": "女"},
            "disability_pre": {"no": "否", "yes": "是"},
            "disability_now": {"no": "否", "yes": "是"},
            "last_status": {"startServer": "服務中",
                            "continue": "服務中",
                            "closed": "結案",
                            "pause": "暫停服務"}
        },
        inplace=True
    )
    # Check if all needed column exist
    errors = []
    for needed_col in [
        "CMSLevel_now", "caseType_now", "sex", "socialWelfare_now"
    ]:
        if any(patient[needed_col].isna()):
            name_list = patient[
                patient[needed_col].isna()]["patient"].astype(str).to_list()
            errors.append(f"{needed_col} missing: {', '.join(name_list)}")
    if len(errors) > 0:
        for error in errors:
            print(error, flush=True)
        raise ReportGenerateError(
            "<br>".join(["核定照顧計畫內缺少必要的欄位資訊：", *errors]))
    sheet1_result = pd.DataFrame()
    for stat in ("_pre", "_now"):
        col_name = {
            "CMSLevel" + stat: "CMSLevel",
            "sex": "sex",
            "caseType" + stat: "caseType",
            "socialWelfare" + stat: "socialWelfare",
        }
        temp_df = patient.groupby(
            list(col_name.keys())
        ).agg({"patient": "nunique"})
        if stat == "_pre":
            temp_df["patient"] = -temp_df["patient"]
        temp_df = temp_df.reset_index().rename(columns=col_name)
        sheet1_result = pd.concat([sheet1_result, temp_df])
    sheet1_result = sheet1_result.groupby(list(col_name.values())).agg({
        "patient": "sum"}).reset_index().rename(columns={"patient": "amount"})
    sheet1_result = pd.melt(
        sheet1_result,
        id_vars=["CMSLevel", "sex", "caseType", "socialWelfare"],
        var_name="table"
    ).pivot_table(
        index=["CMSLevel", "sex", "caseType", "socialWelfare", "table"]
    ).to_dict()["value"]
    sheet2_col = [
        "reason",
        "displayName",
        "sex",
        "age",
        "b_day",
        "last_status",
        "DCHCServiceType",
        "socialWelfareStatus_now",
        "socialWelfareStatus_pre",
        "CMSLevel_now",
        "CMSLevel_pre",
        "socialWelfare_now",
        "socialWelfare_pre",
        "disability_now",
        "disability_pre",
        "caseType_now",
        "caseType_pre",
        "start_at",
        "close_at",
        "BA",
        "GA",
    ]
    for col in sheet2_col:
        if col not in patient.columns:
            patient[col] = None
    sheet2_result = patient[sheet2_col]

    # Filing section
    report_file = BytesIO()
    f_path = path.dirname(path.dirname(path.abspath(__file__)))
    try:
        work_book = load_workbook(
            "pipelines/ds_reportplatform_generator/templates/"
            "hwa_hc_community_diff_template.xlsx"
        )
    except FileNotFoundError:
        work_book = load_workbook(
            path.join(
                f_path, "templates", "hwa_hc_community_diff_template.xlsx"))

    ws1 = work_book["本期服務人數"]
    ws1["C4"] = sheet_title
    for k, v in sheet1_result.items():
        y = cms_coordinate[k[0]] + sex_coordinate[k[1]]
        x = get_column_letter(case_coordinate[k[2]] + social_coordinate[k[3]])
        pos = x + str(y)
        ws1[pos] = v

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = work_book
        writer.sheets = dict((ws.title, ws) for ws in work_book.worksheets)

        # Fill in data to sheet 2
        sheet2_result.to_excel(
            writer,
            sheet_name="服務個案清單",
            header=False,
            index=False,
            startrow=1)

    return report_file
