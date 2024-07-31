"""
ReportName: 桃園居服報表
POC: Shen Chiang
"""

from datetime import datetime, time
from functools import reduce
from io import BytesIO
from os import path

import pandas as pd
from bson.objectid import ObjectId
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pytz import timezone

from linkinpark.app.ds.reportPlatformBackend.utils import (ReportEmptyError,
                                                           ReportGenerateError,
                                                           check_org_type,
                                                           count_age,
                                                           get_nis_data)

UTC = timezone("utc")
TWZ = timezone("Asia/Taipei")


def query_data_of_patients(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function will query the patient data with care plan and service records,
    which will be used in 1st sheet of Taoyuan home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :return: A dataframe to be filled in sheet 4.
    """
    # parameter setting section
    collections = [
        "patients",
        "punchclockrecords",
        "daycareservices",
        "approvedcareplans",
        "transfermanages",
    ]
    conditions = [
        {"organization": org,
         "isDeleted": {"$ne": True},
         "DCHCServiceType": "homeCare",
         "residentialAddressArea": {"$in": ["新屋區", "楊梅區"]}
         },
        {"organization": org,
         "date": {"$gte": start, "$lt": end},
         "serviceStatus": {"$in": ["normal", "unmet"]}},
        {"organization": org},
        {"organization": org,
         "dateOfApproval": {'$lt': end},
         "planType": {"$ne": "termination"}},
        {"organization": org,
         "createdDate": {'$lt': end}},
    ]
    columns = [
        {"birthday": 1,
         "numbering": 1,
         "caseNumber": 1,
         "sex": 1,
         "residentialAddressArea": 1,
         "DCHCServiceType": 1,
         "lastName": 1,
         "firstName": 1,
         "handicapHandbook": 1},
        {"patient": 1, "date": 1, "services": 1},
        {"code": 1},
        {"CMSLevel": 1,
         "dateOfApproval": 1,
         "socialWelfareStatus": 1,
         "socialWelfare": 1,
         "patient": 1,
         "disability": 1},
        {"patient": 1,
         "createdDate": 1,
         "status": 1,
         "reason": 1,
         "firstServerDate": 1},
    ]

    # Querying section
    patient, records, services, care_plan, status = get_nis_data(
        collections, conditions, columns)
    if patient.empty:
        raise ReportEmptyError(
            "無法產製此報表，因無符合條件的個案。"
            "服務類別須為居家服務且居住區域須為新屋區或楊梅區")

    # Preprocess Section
    patient = patient.rename(columns={"_id": "patient"})
    for col in columns[0]:
        if col not in patient.columns:
            patient[col] = None
    patient["age"] = count_age(patient["birthday"], start)
    records = records.explode("services").rename(
        columns={"_id": "record_id"}).reset_index(drop=True)
    records = pd.merge(
        records,
        records["services"].apply(pd.Series),
        left_index=True,
        right_index=True
    )

    services = services.rename(columns={"_id": "service"})
    services["category"] = services["code"].str.slice(0, 2)
    services = services[services["category"].isin(["BA", "GA"])]
    care_plan = care_plan.loc[
        care_plan.sort_values("dateOfApproval").groupby("patient")[
            "dateOfApproval"].idxmax()
    ].reset_index(drop=True)
    care_plan["socialWelfare"] = care_plan["socialWelfare"].astype(str)
    for pattern, replacement in {
        "[": "",
        "'": "",
        "]": "",
        "lowInLaw": "長照低收",
        "lowToMid": "長照中低收",
        "low": "長照低收",
        "normal": "一般戶",
        "nan": "未填寫",
    }.items():
        care_plan["socialWelfare"] = care_plan["socialWelfare"].str.replace(
            pattern, replacement, regex=False)
    status.loc[
        ~status["firstServerDate"].isna(), "createdDate"
    ] = status["firstServerDate"]
    stats_dfs = []
    stats = ("signAwait", "cancelAwait", "startSoon", "startServer", "closed")
    for stat in stats:
        df = status.loc[
            status["status"] == stat, ["patient", "createdDate", "reason"]]
        df = df.loc[df.groupby("patient")["createdDate"].idxmax()]
        stats_dfs.append(df.rename(columns={"createdDate": stat + "_at",
                                            "reason": stat + "_reason"}))
    last_df = status.loc[status.groupby("patient")["createdDate"].idxmax()]
    last_df = last_df[["patient", "status"]].rename(
        columns={"status": "last_status"})
    stats_dfs.append(last_df)

    status = reduce(
        lambda left, right: pd.merge(
            left, right, on=['patient'], how='outer'), stats_dfs)

    # Merging section
    records = pd.merge(records, services, on="service")
    patient = pd.merge(patient, care_plan, "left", on="patient")
    patient = pd.merge(patient, status, "left", on="patient")

    # Computing section
    start_stat = ["signAwait_at", "startSoon_at", "startServer_at"]
    patient["max_start"] = patient[start_stat].max(axis=1)
    patient["reopen"] = False
    patient.loc[patient["max_start"] > patient["closed_at"], "reopen"] = True
    for stat in stats:
        col = stat + "_at"
        patient.loc[patient["reopen"]
                    & (patient[col] < patient["closed_at"]), col] = None
    patient["start_at"] = patient[[
        "signAwait_at", "startSoon_at", "startServer_at"]].min(axis=1)
    for col in ("cancelAwait_at", "closed_at"):
        patient.loc[patient[col] <= patient["start_at"], col] = None
    patient = patient[
        (patient["start_at"].dt.date < end.date())
        & ((patient["closed_at"].dt.date >= start.date())
           | pd.isna(patient["closed_at"]))]
    for row in patient.index:
        dementia, age, status = patient.loc[
            row, ["disability", "age", "socialWelfareStatus"]]
        if 50 <= age and dementia == "yes":
            patient.loc[row, "caseType"] = "50歲以上失智症"
        elif 55 <= age <= 64 and "indigenous" in status:
            patient.loc[row, "caseType"] = "55-64歲原住民"
        elif age <= 64 and "obstacle" in status:
            patient.loc[row, "caseType"] = "64歲以下領身心障礙證明者"
        elif 65 <= age and "obstacle" in status:
            patient.loc[row, "caseType"] = "65歲以上領身心障礙證明者"
        elif 65 <= age:
            patient.loc[row, "caseType"] = "65歲以上老人"
        else:
            patient.loc[row, "caseType"] = "無法判斷"
    amount_df = records.groupby("patient").agg(
        {"record_id": "nunique"}).reset_index().rename(
        columns={"record_id": "amount"}
    )
    records = records.groupby(["patient", "category", "date"]).agg(
        {"record_id": "nunique"}).reset_index()
    records["present"] = (
        records["date"].dt.date.astype(str)
        + "(" + records["record_id"].astype(str) + ")")
    records_dict = {}
    for _id in records["patient"].unique():
        records_dict[_id] = {}
        for cat in ("BA", "GA"):
            date = list(
                records.loc[
                    (records["patient"] == _id)
                    & (records["category"] == cat),
                    "present"])
            records_dict[_id][cat] = sorted(date)
    records = pd.DataFrame.from_dict(
        records_dict, orient="index"
    ).reset_index().rename(columns={"index": "patient"})
    result = pd.merge(patient, records, "inner", on="patient")
    result = pd.merge(result, amount_df, "left", on="patient")
    date_col = [
        "birthday",
        "dateOfApproval",
        "signAwait_at",
        "cancelAwait_at",
        "startSoon_at",
        "startServer_at",
        "closed_at",
        "start_at"
    ]
    for col in date_col:
        result[col] = result[col] + pd.Timedelta(hours=8)
    return result


def query_data_of_users(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function is used to query the user .
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end:  End date of report period.
    :return: A dataframe of user records.
    """
    # Parameter setting section
    user_columns = {
        "displayName": 1,
        "employeeNumber": 1,
        "jobTitle": 1,
        "jobType": 1,
        "staffStatus": 1,
        "sex": 1,
        "resignDate": 1,
    }

    # Querying section
    df = get_nis_data(
        "users",
        {"organization": org,
         "isDeleted": {"$ne": True},
         "jobType": {"$in": ["homeAttendant", "supervisor"]},
         "employDate": {"$lt": end},
         "staffStatus": {"$ne": "resigned"},
         "$or": [{"resignDate": None}, {"resignDate": {"$gte": start}}]},
        user_columns,
    )

    # Preprocess Section
    for col in user_columns:
        if col not in df.columns:
            df[col] = None
    df["resignDate"] = pd.to_datetime(df["resignDate"]) + pd.Timedelta(hours=8)
    df["displayName"] = df["employeeNumber"].fillna("") + " " + df[
        "displayName"]
    df["displayStatus"] = df["staffStatus"]
    df.replace({
        "sex": {"male": "男", "female": "女"},
        "jobType": {"homeAttendant": "居家服務員", "supervisor": "居服督導"},
        "displayStatus": {
            "employed": "在職", "furlough": "留職停薪", "resigned": "離職"
        }
    }, inplace=True)
    df.loc[
        df["staffStatus"] == "resigned", "displayStatus"
    ] += df["resignDate"].dt.strftime("(%Y-%m-%d)")
    return df


def query_data_of_clock_in(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function is used to query the data of clock in hours.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date fo report period.
    :return: A dataframe of clock in records.
    """
    # Parameter setting section
    clock_columns = {
        "_id": 0,
        "patient": 1,
        "clockIn": 1,
        "clockOut": 1,
        "user": 1,
        "staff": 1,
    }

    # Querying section
    df = get_nis_data(
        "punchclockrecords",
        {"organization": org,
         "removed": False,
         "clockIn": {"$gte": start, "$lt": end},
         "serviceStatus": {"$in": ["normal", "unmet"]}},
        clock_columns,
    )

    df["time"] = df["clockOut"].dt.floor("min") - df["clockIn"].dt.floor("min")
    return df


def query_holiday_settings(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function is used to query the data of holiday to count the amount of
    working days in the start and end period.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date fo report period.
    :return: Amount of working days between the period.
    """
    # Querying section
    holiday_df = get_nis_data(
        "festivalsettings",
        {"organization": org,
         "isHoliday": "是",
         "holidayDate": {"$gte": start, "$lt": end}},
        {"_id": 0, "holidayDate": 1}
    )
    week_df = pd.DataFrame(
        pd.Series(pd.date_range(
            start.astimezone(TWZ),
            (end - relativedelta(days=1)).astimezone(TWZ)
        )), columns=["date"]
    )

    # Preprocess Section
    week_df["weekday"] = week_df["date"].dt.weekday
    week_df["workday"] = True
    week_df.loc[week_df["weekday"].isin([5, 6]), "workday"] = False
    if not holiday_df.empty:
        week_df.loc[
            week_df["weekday"].isin(holiday_df["holidayDate"]), "workday"
        ] = False
    return week_df["workday"].sum()


def generate_data_for_table_1(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function is used to generate the data result for the 1st table in
    1st sheet of Taoyuan home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :return: A dictionary of the position and value.
    """
    # Querying section
    df = query_data_of_patients(org, start, end)
    case_type_loc = {
        "65歲以上老人": 6,
        "65歲以上領身心障礙證明者": 8,
        "64歲以下領身心障礙證明者": 10,
        "55-64歲原住民": 12,
        "50歲以上失智症": 14,
    }
    gender_loc = {"male": 0, "female": 1}
    social_status_loc = {
        "長照低收": 3,
        "長照中低收": 21,
        "一般戶": 39,
    }
    cms_level_loc = {
        2: 1,
        3: 3,
        4: 5,
        5: 7,
        6: 9,
        7: 11,
        8: 13
    }
    result = {}
    for case_type, type_row in case_type_loc.items():
        for gender, gender_row in gender_loc.items():
            for status, status_col in social_status_loc.items():
                for cms, level_col in cms_level_loc.items():
                    row = str(type_row + gender_row)
                    amount_col = get_column_letter(status_col + level_col)
                    freq_col = get_column_letter(status_col + level_col + 1)
                    temp_df = df[
                        (df["caseType"] == case_type)
                        & (df["sex"] == gender)
                        & (df["socialWelfare"] == status)
                        & (df["CMSLevel"] == cms)]
                    result[amount_col + row] = len(temp_df)
                    result[freq_col + row] = temp_df["amount"].sum()
    return result


def generate_data_for_table_2(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function is used to generate the data result for the 2nd table in
    1st sheet of Taoyuan home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :return: A dictionary of the position and value.
    """
    # Parameter setting section
    case_type_loc = {
        "65歲以上老人": 77,
        "65歲以上領身心障礙證明者": 79,
        "64歲以下領身心障礙證明者": 81,
        "55-64歲原住民": 83,
        "50歲以上失智症": 85,
    }
    gender_loc = {"male": 0, "female": 1}
    cancel_reason_loc = {
        "無意願": "G",
        "自行照顧": "H",
        "住院中": "I",
        "聘雇外看": "J",
        "其他": "K",
    }
    close_reason_loc = {
        "notWilling": "M",
        "Family": "N",
        "foreignCare": "O",
        "death": "P",
        "other": "Q"
    }

    # Querying section
    df = query_data_of_patients(org, start, end)

    # Preprocess Section
    df = df[((df["start_at"].dt.date >= start.date())
             & (df["start_at"].dt.date < end.date()))
            | (df["last_status"] == "closed")]
    df.loc[
        ~df["cancelAwait_reason"].isin(
            ["無意願", "自行照顧", "住院中", "聘雇外看"]),
        "cancelAwait_reason"
    ] = "其他"
    df.loc[
        ~df["closed_reason"].isin(
            ["notWilling", "Family", "foreignCare", "death"]
        ), "closed_reason"
    ] = "other"

    # Computing section
    result = {}
    for case_type, type_loc in case_type_loc.items():
        for gender, sex_loc in gender_loc.items():
            temp_df = df[(df["caseType"] == case_type) & (df["sex"] == gender)]
            row = str(case_type_loc[case_type] + gender_loc[gender])
            result["D" + row] = len(
                temp_df[(temp_df["start_at"].dt.date >= start.date())
                        & (temp_df["start_at"].dt.date < end.date())]
            )
            result["E" + row] = len(
                temp_df[pd.isna(temp_df["cancelAwait_at"])])
            result["F" + row] = len(
                temp_df[~pd.isna(temp_df["cancelAwait_at"])])
            for reason, reason_loc in cancel_reason_loc.items():
                result[reason_loc + row] = len(
                    temp_df[(temp_df["cancelAwait_reason"] == reason)
                            & (~pd.isna(temp_df["cancelAwait_at"]))]
                )
            for reason, reason_loc in close_reason_loc.items():
                result[reason_loc + row] = len(
                    temp_df[(temp_df["closed_reason"] == reason)
                            & (~pd.isna(temp_df["closed_at"]))]
                )
    return result


def generate_data_for_table_3(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function is used to generate the data result for the 3rd table in
    1st sheet of Taoyuan home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period
    :return: A dictionary of the position and value.
    """
    # Querying section
    df = query_data_of_users(org, start, end)

    # Preprocess Section
    result = {
        "V77": ("居家服務員", "男"),
        "W77": ("居家服務員", "女"),
        "X77": ("居服督導", "男"),
        "Y77": ("居服督導", "女"),
    }
    for loc, condition in result.items():
        result[loc] = len(df[(df["jobType"] == condition[0])
                             & (df["sex"] == condition[1])])

    return result


def generate_data_for_table_6(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function is used to generate the data result for the 6th table in
    1st sheet of Taoyuan home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :return: A dictionary of the position and value.
    """
    # Querying section
    df = generate_data_for_sheet_3(org, start, end)

    # Preprocess Section
    result = {
        "W87": ("全職", "男"),
        "X87": ("全職", "女"),
        "Z87": ("兼職", "男"),
        "AA87": ("兼職", "女"),
    }
    for loc, condition in result.items():
        result[loc] = len(df[(df["類型"] == condition[0])
                             & (df["性別"] == condition[1])])
    return result


def generate_data_for_sheet_2(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function is used to generate the data result for the 2nd sheet (patients
    details) in Taoyuan home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :return: A dataframe to be filled in sheet 2.
    """
    df = query_data_of_patients(org, start, end)
    df["displayName"] = (
        "[" + df["residentialAddressArea"].str.slice(0, 2) + "]"
        + df["numbering"].fillna("") + " "
        + df["lastName"] + df["firstName"])
    replace_symbol = {"[": "", "]": "", "'": ""}
    replace_dict = {
        "DCHCServiceType": {
            **replace_symbol,
            "HCOwnExpense": "自費",
            "HCRespite": "居家喘息",
            "homeCare": "居家服務",
        },
        "socialWelfareStatus": {
            **replace_symbol,
            "general": "一般戶",
            "SocialAssistanceLegalLowIncome": "社會救助法定低收入戶",
            "indigenous": "原住民",
            "lowIncome": "長照低收",
            "middleIncome": "長照中低收",
            "obstacle": "領有身心障礙證明",
            "veteran": "榮民",
        },
        "socialWelfare": {
            **replace_symbol,
            "lowInLaw": "長照低收",
            "lowToMid": "長照中低收",
            "low": "長照低收",
            "normal": "一般戶",
            "nan": "未填寫",
        },
        "BA": {**replace_symbol, "nan": ""},
        "GA": {**replace_symbol, "nan": ""},
    }
    for col, replace in replace_dict.items():
        df[col] = df[col].astype(str)
        for k, v in replace.items():
            df[col] = df[col].str.replace(k, v, regex=True)
    df.replace(
        {
            "sex": {"male": "男", "female": "女"},
            "disability": {"no": "否", "yes": "是"},
            "last_status": {"startServer": "服務中",
                            "continue": "服務中",
                            "closed": "結案",
                            "pause": "暫停服務"},
            "handicapHandbook": {"yes": "是",
                                 "no": "否",
                                 "applying": "申請中"},
            "closed_reason": {"notWilling": "無意願",
                              "Family": "自行照顧",
                              "foreignCare": "聘雇外看",
                              "death": "死亡"}
        },
        inplace=True
    )
    for col in ("birthday", "signAwait_at", "startSoon_at",
                "startServer_at", "closed_at", "cancelAwait_at"):
        df[col] = df[col].dt.tz_localize(UTC)
        df[col] = df[col].dt.tz_convert(TWZ)
        df[col] = df[col].dt.date

    col_names = {
        "displayName": "個案",
        "sex": "性別",
        "age": "年齡",
        "birthday": "生日",
        "last_status": "服務狀態",
        "DCHCServiceType": "服務類別",
        "residentialAddressArea": "居住區域",
        "handicapHandbook": "身心障礙證明/手冊",
        "CMSLevel": "CMS等級",
        "socialWelfare": "長照福利身份",
        "disability": "身心障礙-失智",
        "caseType": "符合資格",
        "signAwait_at": "登記日期",
        "startSoon_at": "安排日期",
        "startServer_at": "開始服務日期",
        "closed_at": "結案日期",
        "closed_reason": "結案原因",
        "cancelAwait_at": "取消候補日期",
        "cancelAwait_reason": "取消原因",
        "BA": "BA碼打卡日期",
        "GA": "GA碼打卡日期",
    }
    df = df[col_names].rename(columns=col_names)
    return df


def generate_data_for_sheet_3(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taoyuan_homecare_community_report. This
    function is used to generate the data result for the 3rd sheet (user
    details) in Taoyuan home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :return: A dataframe to be filled in sheet 3.
    """
    # Querying section
    user_df, clock_df = (query_data_of_users(org, start, end),
                         query_data_of_clock_in(org, start, end))
    work_days = query_holiday_settings(org, start, end)

    # Preprocess Section
    user_df = user_df[user_df["jobType"] == "居家服務員"]
    user_df.rename(columns={"_id": "user"}, inplace=True)
    clock_df = pd.DataFrame(
        clock_df.groupby("staff")["time"].sum().dt.total_seconds() / 60
    ).reset_index()
    clock_df["hour"], clock_df["minute"] = clock_df["time"].divmod(60)
    clock_df["present"] = (
        clock_df["hour"].astype(int).astype(str) + "時"
        + clock_df["minute"].astype(int).astype(str) + "分")
    clock_df.rename(columns={"staff": "user"}, inplace=True)

    # Merging Section
    result = pd.merge(user_df, clock_df, "left", on="user")

    # Computing section
    title = f"本月工作日{work_days}天，專職人員應服務滿{work_days * 8}小時。"
    result["note"] = None
    result["employType"] = "兼職"
    result.loc[result["hour"] >= (work_days * 8), "employType"] = "全職"
    result["present"].fillna("0時0分", inplace=True)
    keep_col = {
        "displayName": "人員",
        "sex": "性別",
        "jobType": "類別",
        "jobTitle": "職稱",
        "employType": "類型",
        "displayStatus": "在職狀況",
        "present": "簽到退時數",
        "note": title,
    }
    result = result[keep_col.keys()].rename(columns=keep_col)

    return result


def taoyuan_homecare_community_report(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a statistic report of home care service
    provided by institution at Taoyuan.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    if not check_org_type(org, "homecare"):
        raise ReportGenerateError("此為居服專屬報表，無法應用於其他機構類型。")

    # parameter setting section
    _ = suffix
    start, end = (TWZ.localize(datetime.combine(start, time())),
                  TWZ.localize(datetime.combine(end, time())))
    query_start, query_end = start.astimezone(UTC), end.astimezone(UTC)

    # Querying section
    sheet_1_result = {}
    # The designer of this report only ask for the content of table 1, 2,
    # 3 and 6. Therefore, there will not have functions for getting the date
    # for table 4 or 5.
    table_1 = generate_data_for_table_1(org, query_start, query_end)
    table_2 = generate_data_for_table_2(org, query_start, query_end)
    table_3 = generate_data_for_table_3(org, query_start, query_end)
    table_6 = generate_data_for_table_6(org, query_start, query_end)
    for table in (table_1, table_2, table_3, table_6):
        sheet_1_result = {**sheet_1_result, **table}
    sheet_2_result = generate_data_for_sheet_2(org, query_start, query_end)
    sheet_3_result = generate_data_for_sheet_3(org, query_start, query_end)

    # Filing section
    report_file = BytesIO()
    f_path = path.dirname(path.dirname(path.abspath(__file__)))
    try:
        work_book = load_workbook(
            "linkinpark/app/ds/reportPlatformBackend/templates/tyn_hc_community_template.xlsx"
        )
    except FileNotFoundError:
        work_book = load_workbook(
            path.join(f_path, "templates", "tyn_hc_community_template.xlsx"))
    work_sheet = work_book["新版111.9"]
    for loc, value in sheet_1_result.items():
        work_sheet[loc] = value

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = work_book
        writer.sheets = dict((ws.title, ws) for ws in work_book.worksheets)
        sheet_2_result.to_excel(
            writer,
            sheet_name="個案清單",
            header=False,
            index=False,
            startrow=1)
        sheet_3_result.to_excel(
            writer,
            sheet_name="人員清單",
            header=False,
            index=False,
            startrow=1)
    return report_file
