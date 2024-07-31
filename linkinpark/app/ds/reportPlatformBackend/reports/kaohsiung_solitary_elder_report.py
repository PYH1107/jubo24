"""
ReportName: 高雄獨老季報
POC: Shen Chiang
"""

from datetime import datetime
from functools import reduce
from io import BytesIO
from os import path

import pandas as pd
from bson.objectid import ObjectId
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

from linkinpark.app.ds.reportPlatformBackend.utils import (
    ReportEmptyError, ReportGenerateError, check_org_type, count_age,
    get_nis_data, preprocess_date)


def kaohsiung_solitary_elder_report(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a statistic report of home care service
    provided by institution at Kaohsiung to solitary elders.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    if not check_org_type(org, "homecare"):
        raise ReportGenerateError("此為居服專屬報表，無法應用於其他機構類型。")

    # parameter setting section
    _ = suffix
    start, end, query_start, query_end = preprocess_date([start, end])
    collection = [
        "patients",
        "servicemanagements",
        "daycareservices",
        "approvedcareplans",
        "transfermanages",
        "organizations",
    ]
    condition = [
        {"organization": org,
         "livingStatus": {"$in": ["alone", "withMate"]},
         "isDeleted": {"$ne": True}},
        {"organization": org,
         "start": {"$gte": query_start, "$lt": query_end},
         "funding": "subsidy"},
        {"organization": org, "fundType": {"$ne": "ownExpense"}},
        {"organization": org,
         "dateOfApproval": {'$lt': query_end}},
        {"organization": org,
         "createdDate": {'$lt': query_end}},
        {"_id": org},
    ]
    columns = [
        {"birthday": 1,
         "numbering": 1,
         "caseNumber": 1,
         "sex": 1,
         "residentialAddress": 1,
         "residentialAddressCity": 1,
         "residentialAddressArea": 1,
         "livingStatus": 1,
         "lastName": 1,
         "firstName": 1},
        {"patient": 1, "start": 1, "service": 1},
        {"code": 1},
        {"dateOfApproval": 1,
         "socialWelfareStatus": 1,
         "patient": 1},
        {"patient": 1, "createdDate": 1, "status": 1},
        {"name": 1},
    ]

    # Querying section
    patient, service, code, plan, status, organization = get_nis_data(
        collection, condition, columns
    )
    if service.empty:
        raise ReportGenerateError("查詢區間內無相關服務紀錄，故無法產製此報表。")

    # Preprocess Section
    patient["name"] = patient["lastName"] + patient["firstName"]
    patient["age"] = count_age(patient["birthday"], query_start)
    patient["sex"].replace({"male": "男", "female": "女"}, inplace=True)
    patient.rename(columns={
        "_id": "patient",
        "residentialAddressCity": "county",
        "residentialAddressArea": "region",
    }, inplace=True)
    for col in ("region", "county"):
        patient[col] = patient[col].fillna("無法判斷")
    patient["display"] = "[" + patient["region"] + "] " + patient["name"]
    patient["solitary"] = patient["livingStatus"].apply(
        lambda x: set(x).issubset(["alone", "withMate"]))
    plan = plan.loc[plan.sort_values("dateOfApproval").groupby("patient")[
        "dateOfApproval"].idxmax()].reset_index(drop=True)
    plan["indigenous"] = plan["socialWelfareStatus"].fillna("").apply(
        list).map(set("indigenous").issubset)
    start_df = status.loc[
        status["status"] == "startServer", ["patient", "createdDate"]].rename(
        columns={"createdDate": "start_at"})
    start_df = start_df.loc[
        start_df.groupby("patient")["start_at"].idxmax()]
    close_df = status.loc[
        status["status"] == "closed", ["patient", "createdDate"]].rename(
        columns={"createdDate": "close_at"})
    close_df = close_df.loc[
        close_df.groupby("patient")["close_at"].idxmax()]
    last_df = status.loc[status.groupby("patient")["createdDate"].idxmax()]
    last_df = last_df[["patient", "status"]].rename(
        columns={"status": "last_status"})
    status = reduce(
        lambda left, right: pd.merge(left, right, on=['patient'], how='outer'),
        [start_df, close_df, last_df]
    )
    status.loc[status["close_at"] <= status["start_at"], "close_at"] = pd.NaT
    status["last_status"] = status["last_status"].replace({
        "startServer": "開始服務",
        "continue": "服務中",
        "branchTransfer": "服務中",
        "closed": "結案",
        "pause": "暫停服務",
    })
    code.rename(columns={"_id": "service"}, inplace=True)
    service_category = {
        "home": ["BA01", "BA02", "BA07", "BA15-1", "BA15-2", "BA23"],
        "meal": ["BA05-1", "BA05-2", "BA16-1", "BA16-2"],
        "medical": ["BA14"]}
    for category, serve_codes in service_category.items():
        code.loc[code["code"].isin(serve_codes), "category"] = category
    service = pd.merge(service, code, on="service")
    service = service[~pd.isna(service["category"])]
    service["start"] = (service["start"] + pd.Timedelta(hours=8)).dt.strftime(
        "%Y-%m-%d")
    records = service.groupby(["patient", "category", "start"]).agg(
        served=("start", "nunique")).reset_index()
    records["show"] = (
        records["start"] + "(" + records["served"].astype(str) + ")")
    records = records.groupby(["patient", "category"])["show"].apply(
        list).reset_index()
    records["served"] = records["show"].apply(len)

    # Merging section
    df = reduce(
        lambda left, right: pd.merge(
            left, right, on=['patient'], how='left'),
        [patient, plan, status, records]
    )

    # Computing section
    df["elder"] = (df["age"] >= 65) | ((df["age"] >= 55) & df["indigenous"])
    df = df[df["solitary"] & df["elder"]]

    for col in ("livingStatus", "socialWelfareStatus"):
        df[col] = df[col].astype(str)
    region_category = pd.api.types.CategoricalDtype(
        categories=[
            "鹽埕區", "鼓山區", "左營區", "楠梓區", "三民區", "新興區",
            "前金區", "苓雅區", "前鎮區", "旗津區", "小港區", "鳳山區",
            "岡山區", "旗山區", "美濃區", "林園區", "大寮區", "大樹區",
            "仁武區", "大社區", "鳥松區", "橋頭區", "燕巢區", "田寮區",
            "阿蓮區", "路竹區", "湖內區", "茄萣區", "永安區", "彌陀區",
            "梓官區", "六龜區", "甲仙區", "杉林區", "內門區", "茂林區",
            "桃源區", "那瑪夏區"
        ], ordered=True)
    df["region"] = df["region"].astype(region_category)
    sheet_1_result = df.groupby(["region", "category", "sex"]).agg({"served": "sum"})
    if sheet_1_result.empty:
        raise ReportEmptyError("查詢區間內查無符合條件之服務紀錄。")
    sheet_1_result = sheet_1_result.reset_index().pivot(
        index="region", columns=["category", "sex"], values="served")
    sheet_1_result.columns = ["_".join(p) for p in sheet_1_result.columns]
    sheet_1_col = [
        "home_男", "home_女", "meal_男", "meal_女", "medical_男", "medical_女"]
    for col in sheet_1_col:
        if col not in sheet_1_result.columns:
            sheet_1_result[col] = None
    sheet_1_result = sheet_1_result[sheet_1_col]

    sheet_2_result = df.pivot(
        index=[
            "display",
            "sex",
            "age",
            "birthday",
            "county",
            "region",
            "livingStatus",
            "last_status",
            "socialWelfareStatus",
            "start_at",
            "close_at"
        ],
        columns="category",
        values="show"
    ).reset_index()
    for col in ("birthday", "start_at", "close_at"):
        sheet_2_result[col] = (sheet_2_result[col] + pd.Timedelta(
            hours=8)).dt.strftime("%Y-%m-%d")
    if float("nan") in sheet_2_result.columns:
        sheet_2_result.drop(columns=float("nan"), inplace=True)
    replace_symbol = {"[": "", "]": "", "'": "", "nan": ""}
    replace_dict = {
        "livingStatus": {
            **replace_symbol,
            "alone": "獨居",
            "family": "與家人同住",
            "institution": "住在機構",
            "withMate": "與配偶住",
            "other": "與其他人同住",
            "Government": "政府補助居住服務"
        },
        "socialWelfareStatus": {
            **replace_symbol,
            "general": "一般戶",
            "middleIncome": "長照中低收",
            "lowIncome": "長照低收",
            "SocialAssistanceLegalLowIncome": "社會救助法定低收入戶",
            "veteran": "榮民",
            "indigenous": "原住民",
            "obstacle": "領有身心障礙證明",
        },
        "home": replace_symbol,
        "meal": replace_symbol,
        "medical": replace_symbol
    }
    for col, replacements in replace_dict.items():
        if col in sheet_2_result.columns:
            for target, replacement in replacements.items():
                sheet_2_result[col] = sheet_2_result[col].astype(str)
                sheet_2_result[col] = sheet_2_result[col].str.replace(
                    target, replacement, regex=False)
        else:
            sheet_2_result[col] = None

    # Filing section
    sheet_1_name = organization.at[0, "name"][:30]
    roc_year = start.year - 1911
    quarter = pd.Timestamp(start).quarter
    sheet_1_title = f"高雄市{roc_year}年第{quarter}季獨居老人服務人次概況"
    sheet_1_subtitle = (f"本期{roc_year}年{start.month}月至"
                        f"{(end - relativedelta(days=1)).month}月  服  務  成  "
                        f"果  (人次)")

    report_file = BytesIO()
    f_path = path.dirname(path.dirname(path.abspath(__file__)))
    try:
        work_book = load_workbook(
            "linkinpark/app/ds/reportPlatformBackend/templates/"
            "khh_hc_solitary_elder_template.xlsx"
        )
    except FileNotFoundError:
        work_book = load_workbook(
            path.join(
                f_path, "templates", "khh_hc_solitary_elder_template.xlsx"))
    work_sheet = work_book["organization"]
    work_sheet.title = sheet_1_name
    work_sheet["A1"], work_sheet["B2"] = sheet_1_title, sheet_1_subtitle

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = work_book
        writer.sheets = dict((ws.title, ws) for ws in work_book.worksheets)

        # Fill in data to sheet 1
        sheet_1_result.to_excel(
            writer,
            sheet_name=sheet_1_name,
            header=False,
            index=False,
            startrow=5,
            startcol=1)

        # Fill in data to sheet 2
        sheet_2_result.to_excel(
            writer,
            sheet_name="個案清單",
            header=False,
            index=False,
            startrow=1)

    return report_file
