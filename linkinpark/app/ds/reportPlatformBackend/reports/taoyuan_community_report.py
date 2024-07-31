"""
ReportName: 桃園社區月報
POC: Shen Chiang
"""

import json
from datetime import datetime
from io import BytesIO
from os import path

import pandas as pd
from bson.objectid import ObjectId
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from linkinpark.app.ds.reportPlatformBackend.utils import (
    check_org_type,
    ReportEmptyError,
    ReportGenerateError,
    count_age,
    get_nis_data,
    preprocess_date)


def taoyuan_community_report(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a statistic report of service provided by
    institution in Taoyuan city.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Use to exclude patient and user that.
    :return: The report generated.
    """

    if not check_org_type(org, "daycare"):
        raise ReportGenerateError("此為日照專屬報表，無法應用於其他機構類型。")

    # parameter setting section
    start, end, query_start, query_end = preprocess_date([start, end])
    exclude_patient, exclude_user = None, None
    if isinstance(suffix, str):
        exclude_dict = json.loads(suffix)
        if "exclude_patient" in exclude_dict:
            exclude_patient = [
                ObjectId(_id) for _id in exclude_dict["exclude_patient"]]
        if "exclude_user" in exclude_dict:
            exclude_user = [
                ObjectId(_id) for _id in exclude_dict["exclude_user"]]

    # Querying condition section
    patient_condition = {
        "organization": org,
        "isDeleted": {"$ne": True},
        "$or": [{"status": {"$ne": "closed"}},
                {"lastModifiedDate": {"$gte": query_start}}]
    }
    if exclude_patient:
        patient_condition["_id"] = {"$nin": exclude_patient}

    care_plan_condition = {
        "organization": org,
        "dateOfApproval": {'$lt': query_end},
    }

    service_plan_condition = {
        "organization": org,
        '$and': [{'start': {'$lt': query_end}},
                 {'start': {'$gte': query_start}}],
    }

    service_condition = {
        "organization": org,
    }

    user_condition = {
        "organization": org,
        "isDeleted": {"$ne": True},
        "staffStatus": "employed",
        "firstName": {"$nin": ["管理者", "管理員"]}
    }
    if exclude_user:
        user_condition["_id"] = {"$nin": exclude_user}

    org_condition = {"_id": org}

    # Querying columns section
    patients_columns = {
        "lastName": 1,
        "firstName": 1,
        "sex": 1,
        "birthday": 1,
        "status": 1,
    }

    care_plan_columns = {
        "_id": 0,
        "socialWelfare": 1,
        "CMSLevel": 1,
        "caseType": 1,
        "createdDate": 1,
        "patient": 1,
        "dateOfApproval": 1,
        "disability": 1,
        "socialWelfareStatus": 1,
    }

    service_plan_columns = {
        "start": 1,
        "patient": 1,
        "service": 1,
    }

    service_columns = {
        "code": 1,
    }

    user_columns = {
        "sex": 1, "employType": 1, "jobType": 1, "lastName": 1, "firstName": 1
    }

    org_columns = {"name": 1}

    # Filling position parameters section
    table_x_coordinate = {"amount": 11, "freq": 20}
    table_y_coordinate = {"amount": 0, "freq": 28}
    cms_y_coordinate = dict((y, 5 + (y - 1) * 3) for y in range(2, 9))
    sex_x_coordinate = {"male": 1, "female": 2}
    sex_y_coordinate = {"male": 1, "female": 2}
    group_x_coordinate = {
        "65above": 7,
        "65obstacle": 11,
        "64obstacle": 15,
        "indigenous": 19,
        "dementia": 23,
    }
    string_trans_dict_x_coordinate = {
        "low": 1, "mid": 2, "general": 3,
    }
    code_y_coordinate = {
        "BD01": 60,
        "BD02": 61,
        "BD03": 62,
        "GA03&GA04": 63,
        "GA06": 64,
        "other": 65
    }
    employment_y_coordinate = {
        "fullTime": 73,
        "partTime": 75,
    }
    job_x_coorfinate = {
        "social-worker-master": 4,
        "social-worker": 6,
        "nurse-staff": 8,
        "nurse-aide": 10,
        "therapist": 12,
        "司機": 14,
        "chef": 16,
        "other": 18,
    }
    case_type_y_coordinate = {
        "selfpay": 60,
        "allowance": 61,
    }

    # Querying section
    (
        patient,
        care_plan,
        service_plan,
        service,
        user,
        organization,
    ) = get_nis_data(
        [
            "patients",
            "approvedcareplans",
            "servicemanagements",
            "daycareservices",
            "users",
            "organizations",
        ],
        [
            patient_condition,
            care_plan_condition,
            service_plan_condition,
            service_condition,
            user_condition,
            org_condition,
        ],
        [
            patients_columns,
            care_plan_columns,
            service_plan_columns,
            service_columns,
            user_columns,
            org_columns,
        ],
    )

    # Rename columns as key for merging
    patient.rename(columns={"_id": "patient"}, inplace=True)
    service.rename(columns={"_id": "service"}, inplace=True)
    service_plan.rename(columns={"_id": "service_id"}, inplace=True)

    if service_plan.empty:
        raise ReportEmptyError("查詢區間內查無服務計畫，因此無法產製此報表。")

    # Transfer to UTC+8 timezone
    time_series = (
        patient["birthday"],
        care_plan["createdDate"],
        care_plan["dateOfApproval"],
        service_plan["start"],
    )
    for series in time_series:
        series += pd.Timedelta(hours=8)

    if any(patient["birthday"].isna()):
        no_bday = patient.loc[patient["birthday"].isna(), "patient"].to_list()
        raise ReportGenerateError(f"無法產製報表，因有住民未填寫生日。{no_bday}")
    patient["age"] = count_age(patient["birthday"], start)

    # Get latest approve care plan of each patient to get it's latest social
    # welfare status.
    care_plan = care_plan.sort_values("createdDate", ascending=False).groupby(
        "patient").head(1)
    string_trans_dict = {
        "normal": "general",
        "lowToMid": "mid",
        "low": "low",
        "lowInLaw": "low",
    }
    care_plan['socialWelfare'] = care_plan['socialWelfare'].apply(
        lambda cell: cell if isinstance(cell, list) else [])
    for status in string_trans_dict:
        care_plan.loc[
            care_plan["socialWelfare"].apply(
                [lambda cell: any([status in cell])]
            )["<lambda>"],
            "string_trans_dict"
        ] = string_trans_dict[status]
    care_plan["dementia"] = care_plan["disability"].replace({
        "yes": True,
        "no": False,
        None: False,
    })
    care_plan['socialWelfareStatus'] = care_plan['socialWelfareStatus'].apply(
        lambda cell: cell if isinstance(cell, list) else [])
    for identity in ("obstacle", "indigenous"):
        care_plan[identity] = False
        care_plan.loc[
            care_plan["socialWelfareStatus"].apply(
                [lambda x: any([identity in x])]
            )["<lambda>"],
            identity
        ] = True

    # Merging section
    result = pd.merge(patient, care_plan, "left", on="patient")
    result = pd.merge(result, service_plan, "right", on="patient")
    result = pd.merge(result, service, "left", on="service")

    # Computing section
    result.loc[result["age"] >= 65, "group"] = "65above"
    result.loc[
        (result["age"] >= 65) & result["obstacle"],
        "group"
    ] = "65obstacle"
    result.loc[
        (result["age"] < 65) & result["obstacle"],
        "group"
    ] = "64obstacle"
    result.loc[
        (55 <= result["age"]) & (result["age"] < 65) & result["indigenous"],
        "group"
    ] = "indigenous"
    result.loc[
        (50 <= result["age"]) & result["dementia"],
        "group"
    ] = "dementia"

    # Count result for table 收托人數統計 and 收托人次統計
    host_code = [
        "BB01", "BB03", "BB05", "BB07", "BB09", "BB11", "BB13",  # full day
        "BB02", "BB04", "BB06", "BB08", "BB10", "BB12", "BB14",  # half day
    ]
    respite_code = ["GA03", "GA04", "GA05", "GA06"]
    host_result = result[result["code"].isin(
        list(set().union(host_code, respite_code))
    )]
    if "caseType" not in host_result.columns:
        host_result["caseType"] = None
    allowance_result = host_result[host_result["caseType"] != "selfpay"]
    allowance_result = allowance_result.groupby(
        ["CMSLevel", "sex", "group", "string_trans_dict"]
    ).agg({"patient": "nunique", "service_id": "nunique"})
    allowance_result = pd.melt(
        allowance_result.reset_index().rename(
            columns={"patient": "amount", "service_id": "freq"}
        ),
        id_vars=["CMSLevel", "sex", "group", "string_trans_dict"],
        var_name="table"
    ).pivot_table(
        index=["CMSLevel", "sex", "group", "string_trans_dict", "table"]
    ).to_dict()["value"]

    # Count result for table 自費、申請補助人數統計
    case_type_result = host_result.copy()
    case_type_result["caseType"].fillna("allowance", inplace=True)
    case_type_result = case_type_result.groupby(["caseType", "sex"]).agg({
        "patient": "nunique"
    }).to_dict()["patient"]

    # Count result for table 服務項目人數 & 服務項目人次
    service_code = ["BD01", "BD02", "BD03", "GA03", "GA04", "GA06"]
    service_result = result[~result["code"].isin(host_code)].copy()
    service_result.loc[
        ~service_result["code"].isin(service_code), "code"
    ] = "other"
    service_result.loc[
        service_result["code"].isin(["GA03", "GA04"]), "code"
    ] = "GA03&GA04"
    service_result = service_result.groupby(
        ["code", "sex"]
    ).agg({"patient": "nunique", "service_id": "nunique"})
    if len(service_result) > 0:
        service_result = pd.melt(
            service_result.reset_index().rename(
                columns={"patient": "amount", "service_id": "freq"}
            ),
            id_vars=["code", "sex"],
            var_name="table"
        ).pivot_table(
            index=["code", "sex", "table"]
        ).to_dict()["value"]
    else:
        service_result = {}

    # Count result for table 工作人員人數統計
    user["jobType"].replace({
        "o-therapist": "therapist",
        "p-therapist": "therapist",
        "native-nurse-aide": "nurse-aide",
        "foreign-nurse-aide": "nurse-aide",
    },
        inplace=True,
    )
    user.dropna(subset=["jobType", "employType"], inplace=True)
    job_type = [
        "social-worker-master",
        "social-worker",
        "nurse-staff",
        "nurse-aide",
        "therapist",
        "司機",
        "chef",
    ]
    user.loc[~user["jobType"].isin(job_type), "jobType"] = "other"
    user_result = user.groupby(
        ["employType", "sex", "jobType"]
    ).size().to_dict()

    # Section for creating a dictionary with filling position and it's value
    write_result = {
        "A1": f'{organization.at[0, "name"]} {start.year - 1911}年'
              f'{start.strftime("%m")}月日間照顧服務統計表'
    }
    # Create the filling section information for table
    # 收托人數統計 and 收托人次統計
    for key, value in allowance_result.items():
        cms, sex, group, status, table = key
        x_coordinate = get_column_letter(
            group_x_coordinate[group]
            + string_trans_dict_x_coordinate[status]
        )
        y_coordinate = str(
            table_y_coordinate[table]
            + cms_y_coordinate[cms]
            + sex_y_coordinate[sex]
        )
        write_result[x_coordinate + y_coordinate] = value

    # Create the filling section information for table 自費、申請補助人數統計
    for key, value in case_type_result.items():
        case_type, sex = key
        x_coordinate = get_column_letter(
            sex_x_coordinate[sex] + 3
        )
        y_coordinate = str(
            case_type_y_coordinate[case_type]
        )
        write_result[x_coordinate + y_coordinate] = value

    # Create the filling section information for table
    # 服務項目人數 & 服務項目人次
    for key, value in service_result.items():
        serv, sex, table = key
        x_coordinate = get_column_letter(
            table_x_coordinate[table]
            + sex_x_coordinate[sex]
        )
        y_coordinate = str(code_y_coordinate[serv])
        write_result[x_coordinate + y_coordinate] = value

    # Create the filling section information for table
    # 工作人員人數
    for key, value in user_result.items():
        employ, sex, job = key
        x_coordinate = get_column_letter(job_x_coorfinate[job])
        y_coordinate = str(
            employment_y_coordinate[employ]
            + sex_y_coordinate[sex]
        )
        write_result[x_coordinate + y_coordinate] = value

    # Filing section
    report_file = BytesIO()
    f_path = path.dirname(path.dirname(path.abspath(__file__)))
    try:
        work_book = load_workbook(
            "pipelines/ds_reportplatform_generator/templates/"
            "ty_community_template.xlsx"
        )
    except FileNotFoundError:
        work_book = load_workbook(
            path.join(f_path, "templates", "ty_community_template.xlsx"))
    work_sheet = work_book["每月統計表"]
    for position, value in write_result.items():
        work_sheet[position] = value
    work_book.save(report_file)

    return report_file
