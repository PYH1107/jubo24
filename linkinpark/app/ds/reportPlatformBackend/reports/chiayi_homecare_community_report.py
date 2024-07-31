"""
ReportName: 嘉義居服長期照顧十年計畫
POC: Shen Chiang
"""

from datetime import datetime
from functools import reduce
from io import BytesIO
from os import path

import pandas as pd
from bson.objectid import ObjectId
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from linkinpark.app.ds.reportPlatformBackend.utils import (ReportEmptyError,
                                                           ReportGenerateError,
                                                           check_org_type,
                                                           count_age,
                                                           get_nis_data,
                                                           preprocess_date)


def chiayi_homecare_community_report(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a statistic report of home care service
    provided by institution at Chiayi.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    if not check_org_type(org, "homecare"):
        raise ReportGenerateError("此為居服專屬報表，無法應用於其他機構類型。")

    _ = suffix
    start, end, query_start, query_end = preprocess_date([start, end])
    collection = [
        "patients",
        "servicemanagements",
        "daycareservices",
        "approvedcareplans",
        "transfermanages",
    ]
    condition = [
        {"organization": org,
         "DCHCServiceType": "homeCare",
         "isDeleted": {"$ne": True}},
        {"organization": org,
         "start": {"$gte": query_start, "$lt": query_end},
         "funding": "subsidy"},
        {"organization": org},
        {"organization": org,
         "dateOfApproval": {'$lt': query_end},
         "planType": {"$ne": "termination"}},
        {"organization": org,
         "createdDate": {'$lt': query_end}},
    ]
    patient_col = {
        "birthday": 1,
        "numbering": 1,
        "caseNumber": 1,
        "sex": 1,
        "residentialAddressArea": 1,
        "DCHCServiceType": 1,
        "lastName": 1,
        "firstName": 1
    }
    columns = [
        patient_col,
        {"patient": 1, "start": 1, "service": 1},
        {"code": 1},
        {"CMSLevel": 1,
         "dateOfApproval": 1,
         "socialWelfareStatus": 1,
         "socialWelfare": 1,
         "patient": 1,
         "disability": 1},
        {"patient": 1, "createdDate": 1, "status": 1},
    ]
    replace_symbol = {"[": "", "]": "", "'": ""}
    cms_coordinate = dict((y, 5 + y * 3) for y in range(2, 9))
    sex_coordinate = {"男": 1, "女": 2}
    case_coordinate = {
        "65歲以上老人（含IADLs失能且獨居之老人）": 7,
        "65歲以上領身心障礙證明者": 11,
        "64歲以下領身心障礙證明者": 15,
        "55-64歲原住民": 19,
        "50歲以上失智症": 23,
    }
    social_coordinate = {"長照低收": 1, "長照中低收": 2, "一般戶": 3, }

    # Querying section
    patient, service, code, care_plan, status = get_nis_data(
        collection, condition, columns
    )

    # Preprocess Section
    if service.empty:
        raise ReportEmptyError("查詢區間內查無相關服務紀錄。")
    errors = []
    sheet_title = (
        f"中華民國{start.year - 1911}年{'上' if start.month <= 6 else '下'}半年"
        f"（{start.month}月至{end.month - 1}月）"
    )

    for col in patient_col:
        if col not in patient.columns:
            patient[col] = None
    patient["name"] = patient["lastName"] + patient["firstName"]
    patient["age"] = count_age(
        patient["birthday"], end - relativedelta(days=1)
    )
    patient["b_day"] = patient["birthday"].dt.strftime("%Y-%m-%d")
    patient["displayName"] = (
        "[" + patient["residentialAddressArea"].fillna("") + "] "
        + patient["numbering"].fillna("") + " "
        + patient["lastName"].fillna("")
        + patient["firstName"].fillna("")
    )
    patient.rename(columns={"_id": "patient"}, inplace=True)

    start_df = status.loc[
        status["status"] == "startServer",
        ["patient", "createdDate"]
    ].rename(columns={"createdDate": "start_at"})
    start_df = start_df.loc[
        start_df.groupby("patient")["start_at"].idxmax()]
    close_df = status.loc[
        status["status"] == "closed",
        ["patient", "createdDate"]
    ].rename(columns={"createdDate": "close_at"})
    close_df = close_df.loc[
        close_df.groupby("patient")["close_at"].idxmax()]
    last_df = status.loc[status.groupby("patient")["createdDate"].idxmax()]
    last_df = last_df[["patient", "status"]].rename(
        columns={"status": "last_status"})
    status = reduce(
        lambda left, right: pd.merge(
            left, right, on=['patient'], how='outer'),
        [start_df, close_df, last_df]
    )
    for col in ("start_at", "close_at"):
        status[col] = status[col].dt.date

    care_plan = pd.merge(patient, care_plan, "left", on="patient")
    care_plan = care_plan.dropna(subset=["dateOfApproval"])
    care_plan = care_plan.loc[
        care_plan.sort_values("dateOfApproval").groupby("patient")[
            "dateOfApproval"].idxmax()
    ].reset_index(drop=True)
    for needed_col in [
        "socialWelfareStatus", "socialWelfare", "CMSLevel", "disability"
    ]:
        if any(care_plan[needed_col].isna()):
            name_list = care_plan[
                care_plan[needed_col].isna()]['displayName'].to_list()
            errors.append(f"{needed_col} missing: {', '.join(name_list)}")
    code["category"] = code["code"].str.slice(0, 2)
    code.rename(columns={"_id": "service"}, inplace=True)
    if len(errors) > 0:
        for error in errors:
            print(error, flush=True)
        raise ReportGenerateError(
            "<br>".join(["核定照顧計畫內缺少必要的欄位資訊：", *errors]))

    # Merging section
    patient = pd.merge(care_plan, status, "left", on="patient")
    service = pd.merge(service, code, "left", on="service")

    # Computing section
    for row in patient.index:
        dementia, age, status = patient.loc[
            row, ["disability", "age", "socialWelfareStatus"]
        ]
        if 50 <= age and dementia == "yes":
            patient.loc[row, "caseType"] = "50歲以上失智症"
        elif 55 <= age <= 64 and "indigenous" in status:
            patient.loc[row, "caseType"] = "55-64歲原住民"
        elif age <= 64 and "obstacle" in status:
            patient.loc[row, "caseType"] = "64歲以下領身心障礙證明者"
        elif 65 <= age and "obstacle" in status:
            patient.loc[row, "caseType"] = "65歲以上領身心障礙證明者"
        elif 65 <= age:
            patient.loc[row, "caseType"] = "65歲以上老人（含IADLs失能且獨居之老人）"
        else:
            patient.loc[row, "caseType"] = "無法判斷"
    service_dict = {}
    for _id in service["patient"].unique():
        service_dict[_id] = {}
        for cat in ("BA", "GA"):
            service_date = list(
                service.loc[
                    (service["patient"] == _id)
                    & (service["category"] == cat),
                    "start"
                ].dt.strftime("%Y-%m-%d").unique())
            service_dict[_id][cat] = sorted(service_date)
    service = pd.DataFrame.from_dict(
        service_dict, orient="index"
    ).reset_index().rename(columns={"index": "patient"})
    patient = pd.merge(patient, service, "left", on="patient")

    replace_dict = {
        "DCHCServiceType": {
            **replace_symbol,
            "HCOwnExpense": "自費",
            "HCRespite": "居家喘息",
            "homeCare": "居家服務",
        },
        "socialWelfareStatus": {
            **replace_symbol,
            "general": "一般戶",
            "SocialAssistanceLegalLowIncome": "社會救助法定低收入戶",
            "indigenous": "原住民",
            "lowIncome": "長照低收",
            "middleIncome": "長照中低收",
            "obstacle": "領有身心障礙證明",
            "veteran": "榮民",
        },
        "socialWelfare": {
            **replace_symbol,
            "lowInLaw": "長照低收",
            "lowToMid": "長照中低收",
            "low": "長照低收",
            "normal": "一般戶",
            "nan": "未填寫",
        },
        "BA": {**replace_symbol, "nan": ""},
        "GA": {**replace_symbol, "nan": ""},
    }
    for col, replace in replace_dict.items():
        patient[col] = patient[col].astype(str)
        for k, v in replace.items():
            patient[col] = patient[col].str.replace(k, v, regex=True)
    patient.replace(
        {
            "sex": {"male": "男", "female": "女"},
            "disability": {"no": "否", "yes": "是"},
            "last_status": {"startServer": "服務中",
                            "continue": "服務中",
                            "closed": "結案",
                            "pause": "暫停服務"}
        },
        inplace=True
    )

    stat_df = patient.groupby(
        ["CMSLevel", "sex", "caseType", "socialWelfare"]
    ).agg({"patient": "nunique"})
    stat_df = pd.melt(
        stat_df.reset_index().rename(
            columns={"patient": "amount"}
        ),
        id_vars=["CMSLevel", "sex", "caseType", "socialWelfare"],
        var_name="table"
    ).pivot_table(
        index=["CMSLevel", "sex", "caseType", "socialWelfare", "table"]
    ).to_dict()["value"]
    detail_df = patient[[
        "displayName",
        "sex",
        "age",
        "b_day",
        "last_status",
        "DCHCServiceType",
        "socialWelfareStatus",
        "CMSLevel",
        "socialWelfare",
        "disability",
        "caseType",
        "start_at",
        "close_at",
        "BA",
        "GA",
    ]]

    # Filing section
    report_file = BytesIO()
    f_path = path.dirname(path.dirname(path.abspath(__file__)))
    try:
        work_book = load_workbook(
            "pipelines/ds_reportplatform_generator/templates/"
            "chy_hc_community_template.xlsx"
        )
    except FileNotFoundError:
        work_book = load_workbook(
            path.join(f_path, "templates", "chy_hc_community_template.xlsx"))

    ws = work_book["10730-04-08-2本期服務人數"]
    ws["C4"] = sheet_title
    for k, v in stat_df.items():
        y = cms_coordinate[k[0]] + sex_coordinate[k[1]]
        x = get_column_letter(case_coordinate[k[2]] + social_coordinate[k[3]])
        pos = x + str(y)
        ws[pos] = v

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = work_book
        writer.sheets = dict((ws.title, ws) for ws in work_book.worksheets)
        detail_df.to_excel(
            writer,
            sheet_name="服務個案清單",
            header=False,
            index=False,
            startrow=1)

    return report_file
