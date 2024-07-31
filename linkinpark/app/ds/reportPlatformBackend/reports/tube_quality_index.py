"""
ReportName: 管路品質指標
POC: Shen Chiang
"""

from datetime import datetime
from io import BytesIO

import pandas as pd
from bson.objectid import ObjectId
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from linkinpark.app.ds.reportPlatformBackend.utils import (
    ReportGenerateError, get_nis_data, preprocess_date, trans_timezone)


def tube_quality_index(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function is develop to generate a customize report for institution
    彰化養護, to monitor the outcome of there tube removing training program.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """
    # Regulation of this report
    if org != ObjectId("60ebc7a55738dd0028cd6bf9"):
        raise ReportGenerateError(
            "此為彰化養護專屬報表，因為院區設定問題、無法應用於其他機構。")

    # Querying section
    _ = suffix
    times = preprocess_date([datetime(start.year, 1, 1), start, end])
    year, start, end, query_year, query_start, query_end = times
    patient_list = get_nis_data(
        "patients",
        {"organization": org, "isDeleted": {"$ne": True}},
        {})["_id"].to_list()
    query_setting = {
        "collection": [
            "tubes",
            "nursingdiagnoses",
            "transfermanages",
            "organizations",
        ],
        "condition": [
            {"organization": org,
             "createdDate": {"$lt": query_end},
             "patient": {"$in": patient_list},
             "$or": [{"finished": False},
                     {"finishedDate": {"$gte": query_year}}],
             "type": {"$in": ["NG", "foley"]}},
            {"organization": org,
             "patient": {"$in": patient_list},
             "title": {"$in": ["鼻胃管移除計畫", "導尿管移除計畫"]},
             "createdDate": {"$lt": query_end},
             "$or": [{"finished": False},
                     {"finishedDate": {"$gte": query_year}}]},
            {"organization": org, "patient": {"$in": patient_list}},
            {"_id": org},
        ],
        "columns": [
            {"_id": 0,
             "finished": 1,
             "createdDate": 1,
             "type": 1,
             "patient": 1,
             "finishedDate": 1,
             "finishedReason": 1},
            {"_id": 0,
             "finished": 1,
             "finishNote": 1,
             "title": 1,
             "patient": 1,
             "createdDate": 1,
             "finishedDate": 1},
            {"_id": 0,
             "branch": 1,
             "patient": 1,
             "createdDate": 1,
             "status": 1,
             "room": 1,
             "bed": 1},
            {"name": 1, "nickName": 1}
        ],
    }
    tube_df, nursing_df, trans_df, org_df = get_nis_data(
        query_setting["collection"],
        query_setting["condition"],
        query_setting["columns"],
    )
    date_df = pd.DataFrame(
        zip([1] * (end - relativedelta(days=1)).month,
            pd.date_range(year, end, freq="m")),
        columns=["key", "month"]
    )

    # Preprocess Section
    date_df["month"] = date_df["month"].dt.strftime("%Y%m")

    for df in (tube_df, nursing_df, trans_df):
        df["key"] = 1
        for col in ("createdDate", "finishedDate"):
            try:
                df[col] = trans_timezone(df[col], 0, 8, ignore_nan=True)
                df[col + "_ym"] = df[col].dt.strftime("%Y%m")
            except KeyError:
                pass
    nursing_df.loc[nursing_df["title"] == "導尿管移除計畫", "type"] = "foley"
    nursing_df.loc[nursing_df["title"] == "鼻胃管移除計畫", "type"] = "NG"
    trans_df["branch_group"] = trans_df["branch"].copy()
    trans_df.loc[
        trans_df["branch_group"].isin(["1C", "1D", "1E"]), "branch_group"
    ] = "康樂家園"
    branches = ("康樂家園", "2C", "2D", "2E", "3C", "3D", "3E", "全中心")
    trans_df['branch_group'] = trans_df['branch_group'].astype("category")
    trans_df['branch_group'] = trans_df['branch_group'].cat.add_categories([
        "全中心"])

    # Check patient's branch for each month
    branch_df = pd.DataFrame()
    for m in range(1, end.month + 1):
        branch_date = datetime(end.year, m, 1)
        temp = trans_df[trans_df["createdDate"] < branch_date].copy()
        keep_index = temp.groupby("patient")["createdDate"].idxmax()
        temp = temp.loc[keep_index]
        temp["month"] = (branch_date - relativedelta(days=1)).strftime("%Y%m")
        branch_df = pd.concat([
            branch_df,
            temp[["month", "patient", "branch_group", "room", "bed", "status"]]
        ])
    # Match the tube and nursing diagnosis to month.
    result = []
    for df in (tube_df, nursing_df):
        df = pd.merge(df, date_df, on="key")
        df = df[
            (df["createdDate_ym"] <= df["month"])
            & ((df["finishedDate_ym"] >= df["month"])
               | pd.isna(df["finishedDate_ym"]))]
        df["finished"] = False
        df.loc[
            df["finishedDate"].dt.strftime("%Y%m") == df["month"], "finished"
        ] = True
        df = df.drop(columns=["key", "createdDate_ym", "finishedDate_ym"])
        result.append(df)
    tube_df, nursing_df = result

    # Merging section
    df = pd.merge(
        tube_df,
        nursing_df,
        how="left",
        on=["patient", "month", "type"],
        suffixes=("_t", "_n")
    )
    df = pd.merge(df, branch_df, on=["patient", "month"], how="left")

    # Computing section

    # Edge Case:
    # Tube record starts before transfermanages newcomer date. This will
    # cause some tube record without the branch and bed room info for that
    # month. Current solution is to fill in the nearliest branch info
    # available.
    no_branch = df[pd.isna(df["branch_group"])].index.tolist()
    no_branch.reverse()
    for i in no_branch:
        if i == df.index.max():
            continue
        elif df.loc[i, "patient"] == df.loc[i + 1, "patient"]:
            branch_info = df.loc[i + 1, ["branch_group", "room", "bed"]]
            df.loc[i, ["branch_group", "room", "bed"]] = branch_info
    temp_df = df[~pd.isna(df["title"])][["type", "patient"]].drop_duplicates()
    temp_df["trained"] = True
    df = pd.merge(df, temp_df, "left", on=["type", "patient"])

    result = {}
    for branch in branches:
        if branch == "全中心":
            temp_df = df
        else:
            temp_df = df[df["branch_group"] == branch]

        count_rules = {
            "keep_tube": ~pd.isna(temp_df["type"]),
            "train_remove": ~pd.isna(temp_df["title"]),
            "success_remove":
                temp_df["trained"] & temp_df["finished_t"]
                & (temp_df["finishedReason"] == "remove"),
            "train_failed":
                temp_df["finishNote"].str.contains("訓練失敗", na=False),
            "keep_train":
                ~pd.isna(temp_df["title"])
                & ~(temp_df["trained"] & temp_df["finished_t"]
                    & (temp_df["finishedReason"] == "remove"))
                & ~temp_df["finishNote"].str.contains("訓練失敗", na=False),
            "discharge":
                temp_df["finished_t"] & (temp_df["status"] == "discharge"),
            "tube_slip":
                temp_df["finished_t"]
                & temp_df["trained"]
                & temp_df["finishedReason"].str.contains("滑脫", na=False),
            "hosp_remove":
                temp_df["finished_t"]
                & (temp_df["finishedReason"] == "hospital"),
        }
        for cat, con in count_rules.items():
            temp_result = temp_df[con].groupby(["type", "month"]).agg(
                value=("patient", "nunique"),
            ).reset_index()
            temp_result = pd.concat(
                [temp_result,
                 temp_df[con].groupby(["type"]).agg(
                     value=("patient", "nunique"),
                 ).reset_index()]
            )
            temp_result["month"].fillna("all", inplace=True)
            temp_result["category"] = cat
            if branch not in result.keys():
                result[branch] = pd.DataFrame()
            result[branch] = pd.concat([result[branch], temp_result])
        temp_df = pd.merge(
            pd.DataFrame(["NG", "foley"], columns=["type"]),
            pd.DataFrame(count_rules.keys(), columns=["category"]),
            how="cross"
        )
        temp_df = temp_df.merge(
            pd.DataFrame(pd.date_range(year, periods=12, freq="m"),
                         columns=["month"]),
            how="cross",
        )
        temp_df["month"] = temp_df["month"].dt.strftime("%Y%m")
        temp_df = pd.merge(
            temp_df, result[branch], "outer", on=["type", "category", "month"]
        )
        temp_df["month"] = temp_df["month"].str.strip().str[-2:] + "月"
        temp_df["month"] = temp_df["month"].replace("ll月", "總計")
        temp_df["category"] = temp_df["category"].astype("category")
        temp_df["category"] = temp_df["category"].cat.reorder_categories(
            count_rules.keys())
        temp_df = temp_df.pivot(
            index=["type", "category"], columns="month", values="value"
        ).reset_index()
        temp_df["type"] = temp_df["type"].replace({
            "NG": "鼻胃管", "foley": "導尿管"
        })
        temp_df["category"] = temp_df["category"].replace({
            "keep_tube": "管路留置人數",
            "train_remove": "收案訓練人數",
            "success_remove": "成功移除",
            "train_failed": "訓練失敗",
            "keep_train": "持續訓練中",
            "discharge": "退住",
            "tube_slip": "滑脫",
            "hosp_remove": "住院中移除",
        })
        result[branch] = temp_df

    # Filing section
    report_file = BytesIO()
    workbook = Workbook()
    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = workbook
        for branch in branches:
            result[branch].to_excel(
                writer,
                sheet_name=branch,
                index=False,
                header=True,
                startrow=2,
                startcol=0,
            )
    workbook.close()

    # Formatting section
    workbook = load_workbook(report_file)
    workbook.remove(workbook.worksheets[0])
    for branch in branches:
        worksheet = workbook[branch]
        worksheet["A1"] = (
            f"{org_df.at[0, 'name']}管路移除品質指標{start.year - 1911}年"
        )
        worksheet["A1"].font = Font(size=16, bold=True)
        worksheet["A3"] = f"{start.year - 1911}年{branch}護理區"
        worksheet["A3"].font = Font(bold=True)
        for cell in ("A1", "A4", "A12"):
            worksheet[cell].alignment = Alignment(
                horizontal="center", vertical="center"
            )
        for cell in ("A1:O1", "A3:B3", "A4:A11", "A12:A19"):
            worksheet.merge_cells(cell)
        for i in range(1, 16):
            for j in (12, 20):
                cell = get_column_letter(i) + str(j)
                worksheet[cell].border = Border(
                    top=Side(border_style="thin", color="FF000000")
                )
        for i in range(4, 20):
            worksheet["P" + str(i)].border = Border(
                left=Side(border_style="thin", color="FF000000")
            )
        worksheet.column_dimensions["A"].width = 10
        worksheet.column_dimensions["B"].width = 15
    report_file = BytesIO()
    workbook.save(report_file)
    workbook.close()

    return report_file
