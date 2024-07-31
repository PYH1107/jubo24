"""
ReportName: 照顧日誌
POC: Shen Chiang
"""

from datetime import datetime
from io import BytesIO

import pandas as pd
from bson.objectid import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from linkinpark.app.ds.reportPlatformBackend.utils import (ReportEmptyError,
                                                           ReportGenerateError,
                                                           check_org_type,
                                                           count_age,
                                                           get_nis_data,
                                                           preprocess_date,
                                                           schema,
                                                           trans_timezone)


def respite_diary(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a report like the care diary on NIS system,
    but specific for respite services.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    if not check_org_type(org, "daycare"):
        raise ReportGenerateError("此為日照專屬報表，無法應用於其他機構類型。")

    # Querying section
    _ = suffix
    start, end, query_start, query_end = preprocess_date([start, end])
    service_management = get_nis_data(
        "servicemanagements",
        {"organization": org,
         "start": {"$gte": query_start,
                   "$lt": query_end}},
        {"_id": 0,
         "patient": 1,
         "start": 1,
         "end": 1,
         "service": 1},
    )
    daycareservices = get_nis_data(
        "daycareservices",
        {"organization": org,
         "code": {"$in": ["GA03", "GA04"]}},
        {"code": 1},
    )
    patients = get_nis_data(
        "patients",
        {"organization": org, "isDeleted": {"$ne": True}},
        {"sex": 1,
         "firstName": 1,
         "lastName": 1,
         "branch": 1,
         "numbering": 1,
         "birthday": 1,
         "dateOfFirstService": 1},
    )
    care_diary_col = {
        "_id": 0,
        "date": 1,
        "patient": 1,
        "individualActivity": 1,
        "individualActivityDuration": 1,
        "individualActivityNote": 1,
        "mobilityStatus": 1,
        "mobilityDevice": 1,
        "mobilityNote": 1,
        "moringMentalStatus": 1,
        "morningEmotionResponse": 1,
        "afternoonMentalStatus": 1,
        "afternoonEmotionResponse": 1,
        "mentalNote": 1,
        "intakeStatus": 1,
        "intakeDevice": 1,
        "foodType": 1,
        "lunchAte": 1,
        "dessertType": 1,
        "dessertAte": 1,
        "intakeNote": 1,
        "drinkingStatus": 1,
        "drinkingDevice": 1,
        "drinkingAmount": 1,
        "drinkingNote": 1,
        "napTime": 1,
        "napNote": 1,
        "urinationStatus": 1,
        "urinationDevice": 1,
        "diaperCost": 1,
        "didUrinated": 1,
        "urinationTimes": 1,
        "urinationNote": 1,
        "defecationStatus": 1,
        "defecationDevice": 1,
        "didDefecated": 1,
        "defecationTimes": 1,
        "defecationNote": 1,
        "toothCleaning": 1,
        "toothCleaningDevice": 1,
        "bodyCleaning": 1,
        "bodyCleaningDevice": 1,
        "clothChanging": 1,
        "hygieneNote": 1,
        "medication": 1,
        "medicineNote": 1,
        "specialEvent": 1,
        "user": 1,
    }
    care_diary = get_nis_data(
        "carediaries",
        {"organization": org,
         "date": {"$gte": query_start,
                  "$lt": query_end}},
        care_diary_col)
    users = get_nis_data(
        "users",
        {"organization": org},
        {"displayName": 1},
    )
    org_name = get_nis_data(
        "organizations",
        {"_id": org},
        {"_id": 0,
         "name": 1},
    )

    # Preprocess Section
    if service_management.empty:
        raise ReportEmptyError("查詢區間內查無相關服務紀錄。")
    for col in care_diary_col:
        if col not in care_diary.columns:
            care_diary[col] = None
    if len(service_management) > 0:
        service_management["start"] = trans_timezone(
            service_management["start"],
            from_utc=0,
            to_utc=8,
        )
        service_management["end"] = trans_timezone(
            service_management["end"],
            from_utc=0,
            to_utc=8,
        )
        service_management["date_merge"] = service_management["start"].dt.date
    if any(patients["birthday"].isna()):
        no_bday = patients.loc[patients["birthday"].isna()]["_id"].to_list()
        raise ReportGenerateError(f"無法產製報表，因有住民未填寫生日。{no_bday}")
    patients["birthday"] = trans_timezone(
        patients["birthday"], from_utc=0, to_utc=8)
    if len(care_diary) > 0:
        care_diary["date"] = trans_timezone(
            care_diary["date"],
            from_utc=0,
            to_utc=8,
        )
        care_diary["date_merge"] = care_diary["date"].dt.date
    daycareservices.loc[daycareservices["code"] == "GA03", "remark"] = "全日"
    daycareservices.loc[daycareservices["code"] == "GA04", "remark"] = "半日"
    patients["name"] = patients["lastName"] + patients["firstName"]
    patients["age"] = count_age(patients["birthday"], end)

    # Merging section
    result = pd.merge(
        daycareservices,
        service_management,
        how="right",
        left_on="_id",
        right_on="service",
    ).drop(columns=["_id", "service"])
    result = pd.merge(
        result,
        patients,
        left_on="patient",
        right_on="_id"
    ).drop(columns=["_id"])
    result = pd.merge(
        result,
        care_diary,
        how="left",
        on=["date_merge", "patient"]
    ).drop(columns=["date_merge", "_id"])
    result = pd.merge(
        result,
        users,
        how="left",
        left_on="user",
        right_on="_id"
    ).drop(columns=["_id", "user"])

    # Computing section
    result.sort_values(["numbering", "start"], inplace=True)
    result = result[~pd.isna(result["code"])].reset_index(drop=True)
    if result.empty:
        raise ReportEmptyError("查詢區間內查無喘息紀錄。")
    result["length"] = (
        result["start"].dt.strftime("%H:%M")
        + "-"
        + result["end"].dt.strftime("%H:%M"))
    result["index_name"] = (
        result["start"].dt.strftime("%m-%d")
        + result["start"].dt.weekday.replace({
            0: "(一)",
            1: "(二)",
            2: "(三)",
            3: "(四)",
            4: "(五)",
            5: "(六)",
            6: "(日)"}))
    result["age"] = count_age(result["birthday"], end)
    result["display_name"] = (
        "姓名："
        + result["branch"]
        + "-"
        + result["numbering"]
        + result["name"]
        + "        性別："
        + result["sex"].replace(schema.gender)
        + "        年齡："
        + result["age"].astype(str) + "\n"
        + "首次服務日期："
        + result["dateOfFirstService"].dt.strftime("%y-%m-%d")
        + "        編號："
        + result["numbering"])
    result.set_index("index_name", drop=True, inplace=True)
    result.drop(
        columns=[
            "branch",
            "numbering",
            "dateOfFirstService",
            "start",
            "end",
            "sex",
            "branch",
            "birthday",
            "lastName",
            "firstName",
            "age"],
        inplace=True
    )
    result = result[[
        "patient",
        "name",
        "display_name",
        "code",
        "length",
        "remark",
        "individualActivity",
        "individualActivityDuration",
        "individualActivityNote",
        "mobilityStatus",
        "mobilityDevice",
        "mobilityNote",
        "moringMentalStatus",
        "morningEmotionResponse",
        "afternoonMentalStatus",
        "afternoonEmotionResponse",
        "mentalNote",
        "intakeStatus",
        "intakeDevice",
        "foodType",
        "lunchAte",
        "dessertType",
        "dessertAte",
        "intakeNote",
        "drinkingStatus",
        "drinkingDevice",
        "drinkingAmount",
        "drinkingNote",
        "napTime",
        "napNote",
        "urinationStatus",
        "urinationDevice",
        "diaperCost",
        "didUrinated",
        "urinationTimes",
        "urinationNote",
        "defecationStatus",
        "defecationDevice",
        "didDefecated",
        "defecationTimes",
        "defecationNote",
        "toothCleaning",
        "toothCleaningDevice",
        "bodyCleaning",
        "bodyCleaningDevice",
        "clothChanging",
        "hygieneNote",
        "medication",
        "medicineNote",
        "specialEvent",
        "displayName",
    ]]
    result.replace(
        {"mobilityStatus": schema.assist_status,
         "moringMentalStatus": schema.general_status,
         "morningEmotionResponse": schema.emotion_status,
         "afternoonMentalStatus": schema.general_status,
         "afternoonEmotionResponse": schema.emotion_status,
         "intakeStatus": schema.assist_status,
         "foodType": schema.food_type,
         "lunchAte": schema.ate_amount,
         "dessertType": schema.food_type,
         "dessertAte": schema.ate_amount,
         "drinkingStatus": schema.assist_status,
         "drinkingAmount": schema.drink_amount,
         "napTime": schema.sleep_time,
         "urinationStatus": schema.assist_status,
         "didUrinated": schema.yes_no_type,
         "defecationStatus": schema.assist_status,
         "didDefecated": schema.yes_no_type,
         "toothCleaning": schema.assist_status,
         "bodyCleaning": schema.assist_status,
         "clothChanging": schema.change_dress,
         "medication": schema.take_drug_method,
         },
        inplace=True
    )
    result.rename(
        columns={
            "code": "使用碼別",
            "length": "服務時間",
            "remark": "備註",
            "individualActivity": "活動項目",
            "individualActivityDuration": "活動時間 分",
            "individualActivityNote": "備註",
            "mobilityStatus": "行動狀況",
            "mobilityDevice": "輔具使用狀況",
            "mobilityNote": "備註",
            "moringMentalStatus": "上午精神",
            "morningEmotionResponse": "上午情緒",
            "afternoonMentalStatus": "下午精神",
            "afternoonEmotionResponse": "下午情緒",
            "mentalNote": "備註",
            "intakeStatus": "進食狀況",
            "intakeDevice": "輔具使用狀況",
            "foodType": "餐食型態",
            "lunchAte": "午餐進食量",
            "dessertType": "點心型態",
            "dessertAte": "點心進食量",
            "intakeNote": "備註",
            "drinkingStatus": "飲水狀況",
            "drinkingDevice": "輔具使用狀況",
            "drinkingAmount": "飲水量",
            "drinkingNote": "備註",
            "napTime": "午休時間",
            "napNote": "備註",
            "urinationStatus": "排尿狀況",
            "urinationDevice": "輔具使用狀況",
            "diaperCost": "尿片用量 片",
            "didUrinated": "有無排尿",
            "urinationTimes": "排尿次數 次",
            "urinationNote": "備註",
            "defecationStatus": "排便狀況",
            "defecationDevice": "輔具使用狀況",
            "didDefecated": "有無排便",
            "defecationTimes": "排便次數 次",
            "defecationNote": "備註",
            "toothCleaning": "口腔清潔",
            "toothCleaningDevice": "輔具使用狀況",
            "bodyCleaning": "身體清潔",
            "bodyCleaningDevice": "輔具使用狀況",
            "clothChanging": "更換衣物",
            "hygieneNote": "備註",
            "medication": "藥物",
            "medicineNote": "備註",
            "specialEvent": "特殊情況",
            "displayName": "紀錄者"
        },
        inplace=True
    )

    # Filing section
    report_file = BytesIO()
    workbook = Workbook()
    worksheet_to_del = workbook.active

    for patient in result["patient"].unique():
        temp = result[result["patient"] == patient].copy()
        temp.fillna("-", inplace=True)
        sheet_name = temp["name"].iloc[0]
        display_name = temp["display_name"].iloc[0]
        temp.drop(columns=["patient", "name", "display_name"], inplace=True)
        df_list = []
        new_temp = temp.copy()
        cut_df = True
        while cut_df:
            if len(new_temp) > 7:
                df_list.append(new_temp[0:7])
                new_temp = new_temp[7:]
            else:
                df_list.append(new_temp)
                cut_df = False
        repeat_times = len(df_list)
        max_row = repeat_times * 49 + 3
        temp = pd.DataFrame()
        for df in df_list:
            temp = pd.concat([temp, df.reset_index().T])
        temp.reset_index(inplace=True)
        temp["index"].replace("index_name", "", inplace=True)
        worksheet = workbook.create_sheet(title=sheet_name)

        # Setting print format
        worksheet.page_margins = PageMargins(
            left=0.25,
            right=0.25,
            top=0.5,
            bottom=0.5,
            header=0.25,
            footer=0.25,
        )
        worksheet.print_title_cols = 'A:B'
        worksheet.print_title_rows = '1:3'
        worksheet.row_dimensions[3].height = 28
        worksheet.print_area = "A1:" + "I" + str(max_row)

        # Setting form header
        worksheet.column_dimensions["A"].width = 8
        for col in range(2, 10):
            col_name = get_column_letter(col)
            worksheet.column_dimensions[col_name].width = 12

        worksheet["A1"] = org_name["name"][0]
        worksheet["A1"].font = Font(size=14, bold=True)
        worksheet["A1"].alignment = Alignment(horizontal="center")
        worksheet["A2"] = "表單"
        worksheet["A2"].font = Font(bold=True)
        worksheet["B2"] = "照顧日誌"
        worksheet["F2"] = "期間"
        worksheet["F2"].font = Font(bold=True)
        worksheet["G2"] = ("從" + start.strftime("%Y-%m-%d")
                           + "到" + end.strftime("%Y-%m-%d"))
        worksheet["A3"] = "個案"
        worksheet["A3"].font = Font(bold=True)
        worksheet["B3"] = display_name
        worksheet["B3"].alignment = Alignment(wrap_text=True)
        for cell in ["A1:I1", "B2:E2", "G2:I2", "B3:I3"]:
            worksheet.merge_cells(cell)

        # Write data to worksheet
        with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
            writer.book = workbook
            writer.sheets = {worksheet.title: worksheet}
            temp.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False,
                startrow=3,
                startcol=1)
            writer.save()

        # Start dynamic formatting
        # Fill background color for columns
        for i in range(4, max_row + 1):
            cell_a = "".join(("A", str(i)))
            cell_b = "".join(("B", str(i)))
            cell_i = "".join(("I", str(i)))
            worksheet[cell_a].font = Font(bold=True)
            worksheet[cell_a].alignment = Alignment(
                wrap_text=True,
                vertical="center",
            )
            worksheet[cell_a].fill = PatternFill(
                "solid",
                fgColor="00F0F0F0",
            )
            worksheet[cell_b].fill = PatternFill(
                "solid",
                fgColor="00F0F0F0",
            )
            for j in range(1, 10):
                col = get_column_letter(i)
                cell = col + str(i)
                orig_size = worksheet[cell].font.size
                if orig_size is None:
                    worksheet[cell].font += Font(size=10)
                else:
                    worksheet[cell].font += Font(size=10 - orig_size)
            worksheet[cell_a].border += Border(left=Side(style='thin'))
            for cell in [cell_b, cell_i]:
                worksheet[cell].border += Border(right=Side(style='thin'))

        # Insert row header
        row_header = {
            5: "服務\n紀錄",
            8: "個別\n活動",
            11: "行動\n功能",
            14: "心理\n狀態",
            19: "進食",
            26: "飲水",
            30: "休息",
            32: "排尿",
            38: "排便",
            43: "衛生",
            49: "服藥",
            51: "特殊\n情況",
            52: "紀錄者",
        }
        for n in range(repeat_times):
            for row in row_header:
                worksheet["A" + str(n * 49 + row)] = row_header[row]

            # Fill background color for row
            for m in range(1, 10):
                letter = get_column_letter(m)
                cell = letter + str(n * 49 + 4)
                worksheet[cell].font += Font(bold=True)
                worksheet[cell].fill = PatternFill("solid", fgColor="00D0D0D0")
                worksheet[letter + str(n * 49 + 4)].border += Border(
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                )
                for row in [
                    4, 7, 10, 13, 18, 25, 29, 31, 37, 42, 48, 50, 51, 52
                ]:
                    worksheet[letter + str(n * 49 + row)].border += Border(
                        bottom=Side(style='thin')
                    )

            # Merge cell for row header
            merge_loc = [
                (5, 7),
                (8, 10),
                (11, 13),
                (14, 18),
                (19, 25),
                (26, 29),
                (30, 31),
                (32, 37),
                (38, 42),
                (43, 48),
                (49, 50)
            ]
            for start_row, end_row in merge_loc:
                worksheet.merge_cells(
                    "A" + str(n * 49 + start_row)
                    + ":A" + str(n * 49 + end_row))

    report_file = BytesIO()
    workbook.remove(worksheet_to_del)
    workbook.save(report_file)
    workbook.close()

    return report_file
