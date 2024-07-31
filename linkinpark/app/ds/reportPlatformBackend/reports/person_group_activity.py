"""
ReportName: 團體活動報表
POC: Shen Chiang
"""

from datetime import datetime
from io import BytesIO

import pandas as pd
from bson.objectid import ObjectId
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from linkinpark.app.ds.reportPlatformBackend.utils import (
    clients_infile, ReportEmptyError, get_nis_data, trans_timezone)


def person_group_activity(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a report for patients who has participate
    group activity events during the query period.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    # Querying section
    _ = suffix
    start_q, end_q = trans_timezone((start, end), from_utc=8, to_utc=0)
    schedule = get_nis_data(
        "smartschedules",
        {"organization": org,
         "category": "groupActivity",
         "startTime": {"$gte": start_q,
                       "$lt": end_q},
         "deleted": False,
         },
        {"_id": 0,
         "startTime": 1,
         "groupActivity": 1,
         "task": 1,
         },
    )
    if schedule.empty:
        raise ReportEmptyError("查詢區間內查無智慧排成紀錄。")
    schedule.rename(columns={"groupActivity": "activity"}, inplace=True)
    activity_list = schedule["activity"].to_list()
    client_list = clients_infile(start, end, [org])["patient"].to_list()

    patient, activity, org_df = get_nis_data(
        ("patients", "groupactivityrecords", "organizations"),
        ({"organization": org,
          "isDeleted": {"$ne": True},
          "_id": {"$in": client_list}},
         {'activity': {"$in": activity_list}},
         {"_id": org},
         ),
        ({"lastName": 1,
          "firstName": 1},
         {"_id": 0,
          "attendant": 1,
          "individualReaction": 1,
          "activity": 1},
         {"_id": 0,
          "name": 1},
         ),
    )
    if activity.empty:
        raise ReportEmptyError("查詢區間內查無相關團體活動紀錄。")

    # Preprocess Section
    org_name = org_df.at[0, "name"]
    patient.rename(columns={"_id": "patient"}, inplace=True)
    for col in ("attendant", "individualReaction", "activity"):
        if col not in activity.columns:
            activity[col] = None
    activity.rename(columns={"attendant": "patient"}, inplace=True)

    # Merging section
    result = pd.merge(schedule, activity, "left", on="activity")
    result = pd.merge(
        result.explode("patient"), patient, "right", on="patient"
    )

    # Computing section
    result = result.reset_index()
    result_attend = result["individualReaction"].apply(
        pd.Series).reset_index()
    result_attend = result_attend.melt(id_vars="index")
    result_attend.dropna(subset=["value"], inplace=True)
    if not result_attend.empty:
        result_attend = pd.merge(
            result_attend,
            result_attend["value"].apply(pd.Series),
            left_index=True,
            right_index=True,
        )
    else:
        result_attend = pd.DataFrame(
            columns=[
                *result_attend.columns,
                *["attendantWill", "progressEngage", "interaction", "emotion"]
            ])

    result = pd.merge(result, result_attend, "left", on="index")
    result = result[
        (result["patient"].apply(str) == result["variable"])
        | pd.isna(result["variable"])]
    result["name"] = (result["lastName"].fillna("")
                      + result["firstName"].fillna(""))
    result["name"] = result["name"].str.replace(r"[\[:\]]", "", regex=True)
    result.drop(
        columns=[
            "index",
            "activity",
            "lastName",
            "firstName",
            "individualReaction",
            "variable",
            "value",
        ],
        inplace=True)
    result.dropna(subset=["name"], inplace=True)
    result.replace({
        "attendantWill": {
            "highWill": 4,
            "needRemind": 3,
            "needLobby": 2,
            "noWill": 1},
        "progressEngage": {
            "engage": 4,
            "needAssist": 3,
            "lessEngage": 2,
            "noEngage": 1},
        "interaction": {
            "activeInteract": 4,
            "needGuide": 3,
            "lessInteract": 2,
            "noInteract": 1},
        "emotion": {
            "joyful": 4,
            "happy": 3,
            "noReaction": 2,
            "agitated": 1}},
        inplace=True)
    result["startTime"] = result["startTime"].dt.date
    result["sum"] = (
        result["attendantWill"]
        + result["progressEngage"]
        + result["interaction"]
        + result["emotion"])
    result["average"] = result["sum"] / 4
    result.sort_values(["name", "startTime"], inplace=True)
    result.rename(
        columns={
            "task": "活動",
            "startTime": "日期",
            "attendantWill": "出席意願",
            "progressEngage": "過程參與",
            "interaction": "他人互動",
            "emotion": "情緒",
            "sum": "總分",
            "average": "平均",
        },
        inplace=True,
    )

    # Deal with multiple patient with the same name
    name_list = result["name"].unique()
    for name in name_list:
        same_name = result.loc[result["name"] == name, "patient"].unique()
        person_with_this_name = len(same_name)
        if person_with_this_name > 1:
            for i in range(person_with_this_name):
                result.loc[
                    result["patient"] == same_name[i], "name"
                ] = name + f"_{i+1}"

    patient_list = result["patient"].unique()

    # Filing section
    report_file = BytesIO()
    workbook = Workbook()
    ws_del = workbook.active

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = workbook

        for patient in patient_list:
            temp_df = result[result["patient"] == patient]
            name = temp_df["name"].unique()[0]
            temp_df = temp_df[[
                "日期",
                "活動",
                "出席意願",
                "過程參與",
                "他人互動",
                "情緒",
                "總分",
                "平均"]]
            temp_df.to_excel(
                writer,
                sheet_name=name,
                index=False,
                header=False,
                startrow=5,
                startcol=0)
            worksheet = workbook[name]
            worksheet["A1"] = org_name
            worksheet['A1'].font = Font(size=18, bold=True)
            worksheet['A1'].alignment = Alignment(
                horizontal='center', vertical='center'
            )
            worksheet["A2"] = "表單"
            worksheet['A2'].font = Font(bold=True)
            worksheet["B2"] = "團體活動"
            worksheet["E2"] = "月份"
            worksheet['E2'].font = Font(bold=True)
            worksheet["F2"] = (
                start.strftime("%Y.%m")
                + "-"
                + (end - relativedelta(days=1)).strftime("%Y.%m"))
            worksheet["A3"] = "姓名"
            worksheet['A3'].font = Font(bold=True)
            worksheet["B3"] = name
            worksheet["A4"] = "活動評值"
            worksheet['A4'].font = Font(bold=True)
            worksheet['A4'].fill = PatternFill("solid", fgColor="00808080")
            worksheet['A5'] = "日期"
            worksheet['B5'] = "活動"
            worksheet['C5'] = "出席意願"
            worksheet['D5'] = "過程參與"
            worksheet['E5'] = "他人互動"
            worksheet['F5'] = "情緒"
            worksheet['G5'] = "總分"
            worksheet['H5'] = "平均"
            r = len(temp_df.dropna(subset=["日期"]))
            s = len(temp_df) + 6
            worksheet[f'A{s}'] = "效益統計"
            worksheet[f'A{s}'].font = Font(bold=True)
            worksheet[f'A{s}'].fill = PatternFill("solid", fgColor="00808080")
            worksheet[f'A{s + 1}'] = "參與次數"
            worksheet[f'B{s + 1}'] = r
            worksheet[f'C{s + 1}'] = "次數"
            worksheet[f'A{s + 2}'] = "評值向度"
            worksheet[f'C{s + 2}'] = "出席意願"
            worksheet[f'D{s + 2}'] = "過程參與"
            worksheet[f'E{s + 2}'] = "他人互動"
            worksheet[f'F{s + 2}'] = "情緒"
            worksheet[f'H{s + 2}'] = "平均"
            worksheet[f'A{s + 3}'] = "每項度滿分4分"
            worksheet[f'A{s + 3}'].alignment = Alignment(wrapText=True)
            if r == 0:
                worksheet[f'C{s + 3}'] = f'-'
                worksheet[f'D{s + 3}'] = f'-'
                worksheet[f'E{s + 3}'] = f'-'
                worksheet[f'F{s + 3}'] = f'-'
                worksheet[f'H{s + 3}'] = f'-'
            else:
                worksheet[f'C{s + 3}'] = f'=AVERAGE(C6:C{s - 1})'
                worksheet[f'D{s + 3}'] = f'=AVERAGE(D6:D{s - 1})'
                worksheet[f'E{s + 3}'] = f'=AVERAGE(E6:E{s - 1})'
                worksheet[f'F{s + 3}'] = f'=AVERAGE(F6:F{s - 1})'
                worksheet[f'H{s + 3}'] = f'=AVERAGE(H6:H{s - 1})'

                worksheet[f'C{s + 3}'].number_format = "#,##0.00"
                worksheet[f'D{s + 3}'].number_format = "#,##0.00"
                worksheet[f'E{s + 3}'].number_format = "#,##0.00"
                worksheet[f'F{s + 3}'].number_format = "#,##0.00"
                worksheet[f'H{s + 3}'].number_format = "#,##0.00"
            worksheet[f'H{s + 4}'] = "Jubo智慧照護平台"
            worksheet[f'H{s + 4}'].alignment = Alignment(horizontal='right')
            worksheet[f'H{s + 4}'].font = Font(color='00C0C0C0')

            for cell in [
                "A1:H1",
                "B2:D2",
                "F2:H2",
                "B3:D3",
                "A4:H4",
                f'A{s}:H{s}',
            ]:
                worksheet.merge_cells(cell)
            worksheet.column_dimensions["B"].width = 15
            for col in ["A", "C", "D", "E", "F", "G", "H"]:
                worksheet.column_dimensions[col].width = 11
        report_file = BytesIO()
        workbook.remove(ws_del)
        workbook.save(report_file)
        workbook.close()

    return report_file
