"""
ReportName: 花蓮獨老季報
POC: Shen Chiang
"""

from datetime import datetime
from functools import reduce
from io import BytesIO
from os import path

import pandas as pd
from bson.objectid import ObjectId
from openpyxl import load_workbook

from linkinpark.app.ds.reportPlatformBackend.utils import (
    ReportGenerateError, check_org_type, count_age, get_nis_data,
    preprocess_date, trans_timezone)


def hualien_solitary_elder_report(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a statistic report of home care service
    provided by institution at Hualien to solitary elders.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    if not check_org_type(org, "homecare"):
        raise ReportGenerateError("此為居服專屬報表，無法應用於其他機構類型。")

    # parameter setting section
    _ = suffix
    start, end, query_start, query_end = preprocess_date([start, end])
    collection = [
        "patients",
        "servicemanagements",
        "daycareservices",
        "approvedcareplans",
        "transfermanages",
    ]
    condition = [
        {"organization": org,
         "livingStatus": {"$in": ["alone", "withMate"]},
         "isDeleted": {"$ne": True}},
        {"organization": org,
         "start": {"$gte": query_start, "$lt": query_end},
         "funding": "subsidy"},
        {"organization": org, "fundType": {"$ne": "ownExpense"}},
        {"organization": org,
         "dateOfApproval": {'$lt': query_end}},
        {"organization": org,
         "createdDate": {'$lt': query_end}},
    ]
    columns = [
        {"birthday": 1,
         "numbering": 1,
         "caseNumber": 1,
         "sex": 1,
         "residentialAddress": 1,
         "residentialAddressCity": 1,
         "residentialAddressArea": 1,
         "livingStatus": 1,
         "lastName": 1,
         "firstName": 1},
        {"patient": 1, "start": 1, "service": 1},
        {"code": 1},
        {"CMSLevel": 1, "dateOfApproval": 1, "socialWelfare": 1, "patient": 1},
        {"patient": 1, "createdDate": 1, "status": 1},
    ]
    county_index = {
        "花蓮市": 11,
        "鳳林鎮": 17,
        "玉里鎮": 23,
        "新城鄉": 29,
        "吉安鄉": 35,
        "壽豐鄉": 41,
        "光復鄉": 47,
        "豐濱鄉": 53,
        "瑞穗鄉": 59,
        "富里鄉": 65,
        "秀林鄉": 71,
        "萬榮鄉": 77,
        "卓溪鄉": 83,
    }
    age_index = {
        "65~69歲": 1,
        "70~74歲": 2,
        "75~79歲": 3,
        "80~84歲": 4,
        "85歲以上": 5,
    }
    number_index = {
        1: "一",
        2: "二",
        3: "三",
        4: "四",
    }

    # Querying section
    patient, service, code, careplan, status = get_nis_data(
        collection, condition, columns
    )
    if service.empty:
        raise ReportGenerateError("查詢區間內無相關服務紀錄，故無法產製此報表。")

    # Preprocess Section
    patient["name"] = patient["lastName"] + patient["firstName"]
    patient["age"] = count_age(patient["birthday"], start)
    patient["age_g"] = pd.cut(
        patient["age"],
        bins=[64, 69, 74, 79, 84, float("inf")],
        labels=["65~69歲", "70~74歲", "75~79歲", "80~84歲", "85歲以上"]
    )
    patient["b_day"] = patient["birthday"].dt.strftime("%Y-%m-%d")
    patient["sex"].replace({"male": "男", "female": "女"}, inplace=True)
    patient.rename(columns={
        "_id": "patient",
        "residentialAddressCity": "county",
        "residentialAddressArea": "region",
    }, inplace=True)
    for col in ("region", "county"):
        patient[col] = patient[col].fillna("無法判斷")
    patient["display"] = "[" + patient["region"] + "] " + patient["name"]
    careplan = careplan.loc[
        careplan.sort_values("dateOfApproval").groupby("patient")[
            "dateOfApproval"].idxmax()
    ].reset_index(drop=True)
    careplan["socialWelfare"] = careplan["socialWelfare"].astype(str)
    start_df = status.loc[
        status["status"] == "startServer",
        ["patient", "createdDate"]
    ].rename(columns={"createdDate": "start_at"})
    start_df = start_df.loc[
        start_df.groupby("patient")["start_at"].idxmax()]
    close_df = status.loc[
        status["status"] == "closed",
        ["patient", "createdDate"]
    ].rename(columns={"createdDate": "close_at"})
    close_df = close_df.loc[
        close_df.groupby("patient")["close_at"].idxmax()]
    last_df = status.loc[status.groupby("patient")["createdDate"].idxmax()]
    last_df = last_df[["patient", "status"]].rename(
        columns={"status": "last_status"})
    status = reduce(
        lambda left, right: pd.merge(
            left, right, on=['patient'], how='outer'),
        [start_df, close_df, last_df]
    )
    status.loc[status["close_at"] <= status["start_at"], "close_at"] = pd.NaT
    code["category"] = code["code"].str.slice(0, 2)
    status["last_status"] = status["last_status"].replace({
        "startServer": "開始服務",
        "continue": "服務中",
        "branchTransfer": "服務中",
        "closed": "結案",
        "pause": "暫停服務",
    })
    code.rename(columns={"_id": "service"}, inplace=True)
    service = pd.merge(service, code, on="service")
    service["start"] = service["start"] + pd.Timedelta(hours=8)
    service = service[service["category"] == "BA"]
    service["start"] = service["start"].dt.strftime("%Y-%m-%d")
    records = service.groupby("patient").agg(served=("start", "nunique"))
    date_list = []
    for patient_id in records.index:
        serve_date = service[
            service["patient"] == patient_id]["start"].unique()
        serve_date = sorted(list(serve_date))
        date_list.append(serve_date)
    records["date"] = date_list
    records["date"] = records["date"].astype(str)
    records.reset_index(inplace=True)
    replace_symbol = {"[": "", "]": "", "'": ""}
    replace_dict = {
        "livingStatus": {
            "df": patient,
            "replacement": {
                **replace_symbol,
                "alone": "獨居",
                "family": "與家人或其他人同住",
                "institution": "住在機構",
                "Government": "政府補助居住服務",
            }
        },
        "socialWelfare": {
            "df": careplan,
            "replacement": {
                **replace_symbol,
                "lowInLaw": "法定低收",
                "lowToMid": "長照中低收",
                "low": "長照低收",
                "normal": "一般戶",
                "nan": "未填寫",
            },
        },
        "date": {
            "df": records,
            "replacement": replace_symbol,
        },
    }

    for col, item in replace_dict.items():
        df = item["df"]
        mapping = item["replacement"]
        df[col] = df[col].astype(str)
        for key, value in mapping.items():
            df[col] = df[col].str.replace(key, value, regex=False)

    # Merging section
    result = reduce(
        lambda left, right: pd.merge(
            left, right, on=['patient'], how='left'),
        [patient, careplan, status, records]
    )

    # Computing section
    result = result[
        (result["last_status"].isin(["開始服務", "服務中", "暫停服務"]))
        | ((result["last_status"] == "結案")
           & (result["close_at"] >= query_start)
           & (result["close_at"] < query_end))]
    result = result[result["age"] >= 65]
    for date_col in ("start_at", "close_at"):
        result[date_col] = trans_timezone(result[date_col], 0, 8, True, True)
    detail = {
        "display": "個案",
        "sex": "性別",
        "age": "年齡",
        "b_day": "生日",
        "county": "縣市",
        "region": "區域",
        "residentialAddress": "居住地址",
        "livingStatus": "居住狀況",
        "last_status": "服務狀態",
        "CMSLevel": "CMS等級",
        "socialWelfare": "長照福利身份",
        "start_at": "開案日期",
        "close_at": "結案日期",
        "served": "服務日數",
        "date": "服務紀錄日期"}
    detail_df = result[detail.keys()].rename(columns=detail)
    stat_df = result.groupby(["region", "age_g"]).agg(unit=("served", "sum"))
    stat_df = stat_df[stat_df["unit"] > 0].reset_index()

    # Filing section
    sheet1_name = (str(pd.Timestamp(start).year - 1911)
                   + "第"
                   + str(pd.Timestamp(start).quarter)
                   + "季")
    sheet2_name = sheet1_name + "清單"
    report_file = BytesIO()
    f_path = path.dirname(path.dirname(path.abspath(__file__)))
    try:
        work_book = load_workbook(
            "pipelines/ds_reportplatform_generator/templates/"
            "hwa_hc_template.xlsx"
        )
    except FileNotFoundError:
        work_book = load_workbook(
            path.join(f_path, "templates", "hwa_hc_template.xlsx"))

    # Formatting section
    ws1, ws2 = work_book["statistics"], work_book["details"]
    header = ws1["A3"].value
    for key, value in {
        "Q": number_index[pd.Timestamp(start).quarter],
        "M1": str(start.month),
        "M2": str(end.month - 1)
    }.items():
        header = header.replace(key, value)
    ws1["A3"] = header
    for row in stat_df.index:
        index_1, index_2, unit = stat_df.loc[row]
        if index_1 not in county_index:
            continue
        fill_location = ("N" + str(county_index[index_1] + age_index[index_2]))
        ws1[fill_location] = unit
    ws1.title, ws2.title = sheet1_name, sheet2_name

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = work_book
        writer.sheets = dict((ws.title, ws) for ws in work_book.worksheets)
        detail_df.to_excel(
            writer,
            sheet_name=sheet2_name,
            header=False,
            index=False,
            startrow=1)

    return report_file
