"""
ReportName: 就診總表
POC: Shen Chiang
"""

from datetime import datetime
from io import BytesIO

import pandas as pd
from bson.objectid import ObjectId
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from linkinpark.app.ds.reportPlatformBackend.utils import (
    ReportEmptyError, get_nis_data, preprocess_date, trans_timezone)


def treat_plan(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a remind list of patients who should revisit
    the doctor or to renew their drugs.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    # Querying section
    _ = suffix
    start, end, query_start, query_end = preprocess_date([start, end])
    treatment_record_col_dict = {
        "_id": 1,
        "visitDate": 1,
        "hospital": 1,
        "department": 1,
        "doctor": 1,
        "note": 1,
    }
    treatment_record = get_nis_data(
        "treatmentrecords",
        {"organization": org,
         "visitDate": {"$gte": query_start,
                       "$lt": query_end}},
        treatment_record_col_dict,
    )
    treatment_tracking_col_dict = {
        "_id": 1,
        "patient": 1,
        "treatmentRecord": 1,
        "hospital": 1,
        "department": 1,
        "visitDate": 1,
        "hospitalized": 1,
        "doctor": 1,
        "note": 1,
    }
    treatment_tracking = get_nis_data(
        "treatmentrecordtrackings",
        {"organization": org,
         "visitDate": {"$gte": query_start,
                       "$lt": query_end}},
        treatment_tracking_col_dict,
    )
    smart_schedules_col_dict = {
        "_id": 0,
        "subCategory": 1,
        "patientSelected": 1,
        "treatmentRecord": 1,
    }
    smart_schedules = get_nis_data(
        "smartschedules",
        {"organization": org,
         "startTime": {"$gte": query_start,
                       "$lt": query_end},
         "category": "treatmentRecord"},
        smart_schedules_col_dict,
    )
    drug_plan_col_dict = {
        "drugName": 1,
        "hospital": 1,
        "department": 1,
        "patient": 1,
        "recur": 1,
        "endDate": 1,
        "note": 1,
    }
    drug_plan = get_nis_data(
        "mars",
        {"organization": org,
         "endDate": {"$gte": query_start,
                     "$lt": query_end}},
        drug_plan_col_dict,
    )
    patient = get_nis_data(
        "patients",
        {"organization": org, "isDeleted": {"$ne": True}},
        {"firstName": 1, "lastName": 1, "branch": 1, "bed": 1, "room": 1},
    )
    report_title = get_nis_data(
        "organizations",
        {"_id": org},
        {"_id": 0, "name": 1}
    )

    # Preprocess Section
    for key in treatment_record_col_dict:
        if key not in treatment_record.columns:
            treatment_record[key] = None
    for key in treatment_tracking_col_dict:
        if key not in treatment_tracking.columns:
            treatment_tracking[key] = None
    for key in drug_plan_col_dict:
        if key not in drug_plan.columns:
            drug_plan[key] = None
    if smart_schedules.empty:
        raise ReportEmptyError(f"無法產製報表，因為智慧排程內查無就診提醒紀錄。")
    smart_schedules = (
        smart_schedules.explode("patientSelected")
        .rename(columns={"patientSelected": "patient",
                         "subCategory": "type"})
    )
    treatment_tracking.rename(columns={"hospitalized": "type"},
                              inplace=True)
    track_type = []
    for cell in treatment_tracking["type"]:
        cell = pd.Series(cell)
        cell.replace({"nextVisit": "回診",
                      "hospTransfer": "一般住院",
                      "unplannedHosp": "非計畫住院"},
                     inplace=True)
        cell = cell.to_string(index=False)
        track_type.append(cell)
    treatment_tracking["type"] = track_type

    if len(drug_plan) > 0:
        drug_plan_split_department = \
            drug_plan["department"].str.split(r"([科]\d?\.?)", expand=True)
        if len(drug_plan_split_department) == 3:
            drug_plan["department"] = (drug_plan_split_department[0]
                                       + drug_plan_split_department[1])
            drug_plan["doctor"] = drug_plan_split_department[2]
        else:
            drug_plan["doctor"] = None
        drug_plan["endDate"] = trans_timezone(
            drug_plan["endDate"], from_utc=0, to_utc=8, date_only=True)
    else:
        drug_plan["department"], drug_plan["doctor"], drug_plan["endDate"] = (
            None, None, None
        )

    patient["name"] = patient["lastName"] + patient["firstName"]
    patient["room_bed"] = patient["room"] + "-" + patient["bed"]

    # Merging section
    remain_column = [
        "room_bed",
        "name",
        "hospital",
        "department",
        "doctor",
        "type",
        "visitDate",
        "note",
    ]
    treatment_record = pd.merge(
        treatment_record,
        smart_schedules,
        how="left",
        left_on="_id",
        right_on="treatmentRecord"
    )
    treatment_record = pd.merge(
        treatment_record,
        patient,
        left_on="patient",
        right_on="_id"
    )
    treatment_tracking = pd.merge(
        treatment_tracking,
        patient,
        left_on="patient",
        right_on="_id"
    )
    treatment_plan = pd.concat(
        [treatment_record[remain_column],
         treatment_tracking[remain_column]]
    )
    drug_plan = pd.merge(
        drug_plan,
        patient,
        left_on="patient",
        right_on="_id"
    )

    # Computing section
    cus_fill_columns = ["visit", "number", "attachment", "escort", "contact"]
    for col in cus_fill_columns:
        treatment_plan[col] = None

    treatment_plan_col_order = [
        "visit",
        "number",
        "room_bed",
        "name",
        "hospital",
        "department",
        "doctor",
        "type",
        "visitDate",
        "note",
        "attachment",
        "escort",
        "contact",
    ]
    treatment_plan = treatment_plan[treatment_plan_col_order]
    treatment_plan["visitDate"] = trans_timezone(
        treatment_plan["visitDate"], from_utc=0, to_utc=8, date_only=True
    )
    drug_plan_col_order = [
        "room_bed",
        "name",
        "hospital",
        "department",
        "doctor",
        "recur",
        "endDate",
        "note",
    ]
    drug_plan = drug_plan[drug_plan_col_order]
    drug_plan = drug_plan.drop_duplicates()

    # Formatting section
    treatment_plan = treatment_plan.sort_values(["visitDate", "room_bed"])
    treatment_plan = treatment_plan.rename(
        columns={"visit": "親自看診\n(v)",
                 "number": "預約號\n(當天加掛新增)",
                 "room_bed": "房號",
                 "name": "姓名",
                 "hospital": "醫院",
                 "department": "科別",
                 "doctor": "醫生",
                 "type": "類型",
                 "visitDate": "就診時間",
                 "note": "備註事項",
                 "attachment": "檢附看診資料",
                 "escort": "家屬陪同\n（有打v）",
                 "contact": "家屬姓名與聯絡電話",
                 }
    )
    drug_plan = drug_plan.replace(
        {"recur": {"temp": "臨時用藥", "regular": "常規用藥"}}
    )
    drug_plan = drug_plan.sort_values(["endDate", "room_bed"])
    drug_plan = drug_plan.rename(
        columns={"room_bed": "房號",
                 "name": "姓名",
                 "hospital": "醫院",
                 "department": "科別",
                 "doctor": "醫生",
                 "endDate": "藥物到期日",
                 "recur": "類型",
                 "note": "備註事項",
                 }
    )

    # Save section
    section_at = len(treatment_plan) + 4
    report_file = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "就診總表"

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = workbook
        writer.sheets = dict(
            (worksheet.title, ws) for ws in workbook.worksheets
        )
        treatment_plan.to_excel(
            writer,
            sheet_name="就診總表",
            index=False,
            startrow=0,
            startcol=0
        )
        drug_plan.to_excel(
            writer,
            sheet_name="就診總表",
            index=False,
            startrow=section_at,
            startcol=2
        )
        sheet_col_width = {
            "A": 10,
            "B": 15,
            "C": 12,
            "D": 15,
            "E": 45,
            "F": 25,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 20,
        }
        sheet_header_cell = []
        for col in sheet_col_width:
            worksheet.column_dimensions[col].width = sheet_col_width[col]
            first_section_header = col + "1"
            second_section_header = col + str(section_at + 1)
            sheet_header_cell.append(first_section_header)
            sheet_header_cell.append(second_section_header)

        for cell in sheet_header_cell:
            worksheet[cell].fill = PatternFill(
                "solid",
                fgColor="00969696",
            )
            worksheet[cell].font = Font(bold=True)
            worksheet[cell].alignment = Alignment(
                wrap_text=True,
                horizontal="center",
                vertical="top",
            )
        for i in range(4):
            worksheet.insert_rows(1)

        worksheet["A1"] = report_title.at[0, "name"] + "－就診總表"
        worksheet["A1"].font = Font(size=18, bold=True)
        worksheet["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        worksheet.merge_cells("A1:M1")
        worksheet["A3"] = "日期"
        worksheet["B3"] = "".join(
            (str(start), " - ", str(end - relativedelta(days=1))))

    return report_file
