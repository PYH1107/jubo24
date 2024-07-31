"""
ReportName: 跨專業照會
POC: Shen Chiang
"""

from datetime import datetime
from io import BytesIO

import pandas as pd
from bson.objectid import ObjectId
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from linkinpark.app.ds.reportPlatformBackend.utils import (
    ReportEmptyError, clients_infile, get_nis_data, preprocess_date)


def profession_referral(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a report of the communication record of
    inter professionals.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """
    # Querying section
    _ = suffix

    start, end, query_start, query_end = preprocess_date([start, end])
    client_list = clients_infile(start, end, org)["patient"].to_list()
    trans_df, inter_df, patient_df, user_df, org_df = get_nis_data(
        ["transfermanages",
         "interprofessionalnotes",
         "patients",
         "users",
         "organizations"],
        [{"organization": org,
          "createdDate": {"$lt": query_end},
          "status": {"$nin": ["reservation", "unpresented"]},
          "patient": {"$in": client_list}},
         {"organization": org,
          "createdDate": {"$gte": query_start,
                          "$lt": query_end},
          "patient": {"$in": client_list}},
         {"organization": org,
          "isDeleted": {"$ne": True},
          "_id": {"$in": client_list}},
         {"organization": org, "isDeleted": {"$ne": True}},
         {"_id": org}],
        [{"_id": False,
          "patient": True,
          "branch": True,
          "room": True,
          "bed": True,
          "createdDate": True,
          "status": True},
         {"target": True,
          "createdDate": True,
          "question": True,
          "patient": True,
          "user": True},
         {"lastName": True,
          "firstName": True},
         {"lastName": True,
          "firstName": True},
         {"name": True,
          "branch": True}]
    )

    if inter_df.empty:
        raise ReportEmptyError("查詢區間內查無相關跨專業照會紀錄。")
    inter_id_list = inter_df["_id"].to_list()
    reply_df = get_nis_data(
        "interprofessionalnotereplies",
        {"organization": org,
         "patient": {"$in": client_list},
         "interprofessionalNote": {"$in": inter_id_list}},
        {"_id": False,
         "interprofessionalNote": True,
         "target": True,
         "organization": True,
         "patient": True,
         "createdDate": True,
         "reply": True}
    )

    # Preprocess Section
    for df in (inter_df, reply_df):
        if "createdDate" in df.columns:
            df["createdDate"] += pd.Timedelta(hours=8)

    trans_df = trans_df.loc[
        trans_df.groupby("patient")["createdDate"].idxmax()
    ]

    inter_df.rename(
        columns={"_id": "interprofessionalNote",
                 "createdDate": "referDate"},
        inplace=True)
    inter_df = inter_df.explode("target")
    patient_df.rename(columns={"_id": "patient"}, inplace=True)

    # This section is used to avoid key error when the reply dataframe is
    # empty.
    if reply_df.empty:
        reply_df = pd.DataFrame(
            columns=[
                "interprofessionalNote",
                "target",
                "organization",
                "patient",
                "createdDate",
                "reply",
            ]
        )
    reply_df.rename(columns={"createdDate": "replyDate"}, inplace=True)
    user_df.rename(columns={"_id": "user"}, inplace=True)
    user_df["user_name"] = user_df["lastName"] + user_df["firstName"]
    user_df.drop(columns=["lastName", "firstName"], inplace=True)

    # Merging section
    main_df = pd.merge(inter_df, user_df, "left", on="user")
    main_df = pd.merge(main_df, trans_df, "left", on="patient")
    main_df = pd.merge(main_df, patient_df, "left", on="patient")
    main_df = pd.merge(
        main_df,
        reply_df,
        "left",
        on=["interprofessionalNote", "target", "patient"]
    )

    # Computing section
    for col in ("branch", "room", "bed"):
        if col not in main_df.columns:
            main_df[col] = None
    main_df["display_name"] = (
        main_df["branch"].fillna("") + " "
        + main_df["room"].fillna("") + main_df["bed"].fillna("")
        + main_df["lastName"].fillna("") + main_df["firstName"].fillna(""))
    stat_df = main_df.groupby(["branch", "target"]).size().reset_index()

    # This section is used to deal with branches without inter profession
    # records, but still wish to show number zero of that branch on the report.
    for branch in org_df.at[0, "branch"]:
        if branch not in stat_df["branch"].to_list():
            empty_branch = pd.DataFrame(
                [[branch, "social-worker", 0]],
                columns=["branch", "target", 0]
            )
            stat_df = pd.concat(
                [stat_df, empty_branch],
                ignore_index=True
            )

    stat_df = pd.pivot(stat_df, index="branch", columns="target", values=0)

    detail_df = main_df.copy()
    detail_df["reply"] = detail_df["reply"].fillna("尚未回覆")
    detail_df = detail_df[[
        "referDate",
        "display_name",
        "user_name",
        "question",
        "target",
        "reply",
    ]]
    detail_df = pd.pivot(
        detail_df.drop_duplicates(),
        index=["referDate", "display_name", "user_name", "question"],
        columns="target",
        values="reply"
    ).reset_index()

    target_dict = {
        "social-worker": "社工",
        "pharmacist": "藥師",
        "dietitian": "營養師",
        "rehabilitation": "物治",
        "nurse-practitioner": "護理",
        "care": "照服",
        "other": "其他",
        # On 2022-05-19, the customer mentioned that the following
        # profession should be remove from the report since there is no full
        # time employee for those profession and no one will be replying in
        # the system.
        # "OT": "職治",
        # "doctor": "醫師",
        # "counselor": "心理",
    }
    for col in target_dict.keys():
        if col not in stat_df.columns:
            stat_df[col] = None
        if col not in detail_df.columns:
            detail_df[col] = None

    stat_df.rename(columns=target_dict, inplace=True)
    stat_df.fillna(0, inplace=True)
    stat_df = stat_df[target_dict.values()].T

    detail_df["referDate"] = detail_df["referDate"].dt.date
    rename_dict = {
        "referDate": "照會日期",
        "display_name": "住民",
        "user_name": "照會者",
        "question": "照會問題",
    }
    for key, value in target_dict.items():
        rename_dict[key] = value + "回覆"
    detail_df.rename(columns=rename_dict, inplace=True)

    # Filing section
    report_file = BytesIO()
    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title = "被照會職種統計"
    workbook.create_sheet("照會及回覆明細")
    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        stat_df.to_excel(
            writer,
            sheet_name="被照會職種統計",
            index=True,
            header=True,
            startrow=2,
            startcol=0)
        detail_df.to_excel(
            writer,
            sheet_name="照會及回覆明細",
            index=False,
            header=True,
            startrow=2,
            startcol=0)
    workbook.close()

    # Formatting section
    workbook = load_workbook(report_file)
    worksheet1 = workbook["被照會職種統計"]
    title = (
        f"{org_df.at[0, 'name']}跨專業照會月報表{start.year - 1911}年"
        f"{start.month}月"
    )
    worksheet1["A1"] = title
    worksheet1["A1"].font = Font(size=16, bold=True)
    worksheet1["A3"] = None

    worksheet2 = workbook["照會及回覆明細"]
    worksheet2["A1"] = title
    worksheet2["A1"].font = Font(size=16, bold=True)
    sheet2_col_width = {
        "A": 12,
        "B": 30,
        "C": 10,
    }
    for letter in ("D", "E", "F", "G", "H", "I", "J", "K"):
        sheet2_col_width[letter] = 40
    for col, width in sheet2_col_width.items():
        worksheet2.column_dimensions[col].width = width
    cell_to_wrap = []
    for col in sheet2_col_width.keys():
        for row in range(detail_df.shape[0]):
            cell_to_wrap.append(col + str(row + 4))
    for cell in cell_to_wrap:
        worksheet2[cell].alignment = Alignment(wrapText=True, vertical="top")
    report_file = BytesIO()
    workbook.save(report_file)
    workbook.close()

    return report_file
