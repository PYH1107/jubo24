"""
ReportName: 智慧排程活動記錄表
POC: Shen Chiang
"""

import re
from datetime import datetime
from io import BytesIO

import PIL
import pandas as pd
import requests
from bson.objectid import ObjectId
from openpyxl import Workbook, drawing
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.page import PageMargins

from linkinpark.app.ds.reportPlatformBackend.utils import (ReportEmptyError,
                                                           get_nis_data,
                                                           preprocess_date,
                                                           trans_timezone)


def id_to_name(id_list: (tuple, list),
               patient: pd.DataFrame
               ) -> list:
    if not hasattr(id_list, "__iter__"):
        id_list = [id_list]
    return list(patient[patient["_id"].isin(id_list)]["name"])


def smart_schedule_activity(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function origins from familycalendars.py's function
    activity_records. Since the module familycalendars has been merged to
    smart schedule, the origin function is no more activate. Hence this new
    function if built to fulfill the need for activity records
    with photo after familycalendars was deactivated.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    # Querying section
    _ = suffix
    start, end, query_start, query_end = preprocess_date([start, end])
    schedule, patient, org_file = get_nis_data(
        ["smartschedules", "patients", "organizations"],
        [{"organization": org,
          "category": "activity",
          "deleted": False,
          "startTime": {"$lt": query_end},
          "endTime": {"$gte": query_start}},
         {"organization": org},
         {"_id": org}],
        [{"task": 1,
          "startTime": 1,
          "endTime": 1},
         {"numbering": 1,
          "firstName": 1,
          "lastName": 1},
         {"name": 1}]
    )
    if schedule.empty:
        raise ReportEmptyError("查詢區間內查無相關智慧排程。")

    record = get_nis_data(
        "smartschedulerecords",
        {"schedule": {"$in": schedule["_id"].to_list()}},
        {"photo": 1, "schedule": 1, "description": 1, "attendant": 1}
    )
    if record.empty:
        raise ReportEmptyError("查詢區間內查無相關智慧排成紀錄。")
    for col in ("photo", "schedule", "description", "attendant"):
        if col not in record.columns:
            record[col] = None

    # Preprocess Section
    org_name = org_file["name"][0]
    patient = patient.fillna("")
    patient["name"] = (
        patient["numbering"] + patient["lastName"] + patient["firstName"])
    record["patient"] = record["attendant"].apply(id_to_name, patient=patient)
    time_col = ["startTime", "endTime"]
    for col in time_col:
        schedule[col] = trans_timezone(schedule[col], 0, 8)
    schedule = schedule.sort_values(time_col)
    schedule.rename(columns={"_id": "schedule"}, inplace=True)

    # Merging section
    schedule = pd.merge(schedule, record, "left", on="schedule")

    # Computing section
    # No data need to be computed in this report

    wb = Workbook()
    ws_del = wb.active
    pages = len(schedule)
    pages_list = []
    photo_location = []
    for number in range(10, 50, 20):
        for letter in ["A", "C"]:
            photo_location.append(letter + str(number))

    for i in range(pages):
        sheet_name = (schedule["startTime"][i].strftime("%m%d")
                      + schedule["task"][i])
        if len(sheet_name) > 30:
            sheet_name = sheet_name[:30]
        sheet_name = re.sub(r'[^\w]', r'', sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        ws.page_margins = PageMargins(
            left=0.25,
            right=0.25,
            top=0.75,
            bottom=0.75,
            header=0.5,
            footer=0.5,
        )
        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 27
        cells_to_merge = ["A4:D4", "B5:D5", "B6:D6", "B7:D7", "B8:D8", "A9:D9"]
        for cell in cells_to_merge:
            ws.merge_cells(cell)
        pages_list.append(ws)

        pages_list[i]["A1"] = org_name
        a1 = pages_list[i]["A1"]
        a1.font = Font(size=16, bold=True)

        pages_list[i]["A2"] = "表單"
        a2 = pages_list[i]["A2"]
        a2.font = Font(bold=True)

        pages_list[i]["B2"] = "活動紀錄"

        pages_list[i]["C2"] = "日期"
        c2 = pages_list[i]["C2"]
        c2.font = Font(bold=True)

        pages_list[i]["D2"] = schedule["startTime"][i].strftime("%Y-%m-%d")

        pages_list[i]["A4"] = "活動規劃"
        a4 = pages_list[i]["A4"]
        a4.font = Font(bold=True)
        a4.fill = PatternFill("solid", fgColor="00C0C0C0")

        pages_list[i]["A5"] = "活動名稱"
        pages_list[i]["B5"] = schedule["task"][i]
        pages_list[i]["A6"] = "活動參與人員"

        pages_list[i]["B6"] = str(schedule["patient"][i])
        a6 = pages_list[i]["B6"]
        a6.alignment = Alignment(wrap_text=True)

        pages_list[i]["A7"] = "活動時間"
        pages_list[i]["B7"] = (
            schedule["startTime"][i].strftime("%H:%M")
            + "-"
            + schedule["endTime"][i].strftime("%H:%M"))
        pages_list[i]["A8"] = "活動描述"

        pages_list[i]["B8"] = schedule["description"][i]
        b8 = pages_list[i]["B8"]
        b8.alignment = Alignment(wrap_text=True)

        pages_list[i]["A9"] = "活動照片"
        a9 = pages_list[i]["A9"]
        a9.font = Font(bold=True)
        a9.fill = PatternFill("solid", fgColor="00C0C0C0")

        try:
            photo_amount = len(schedule["photo"][i])
        except TypeError:
            photo_amount = 0
        photo_amount = min(photo_amount, 4)
        if photo_amount == 0:
            continue
        for n in range(photo_amount):
            url = schedule["photo"][i][n]
            response = requests.get(url)
            fp = BytesIO(response.content)
            try:
                pil_img = PIL.Image.open(fp)
            except PIL.UnidentifiedImageError:
                continue

            img_width, img_height = pil_img.size
            if img_width > 350 or img_height > 350:
                max_length = max(img_width, img_height)
                ratio = 350 / max_length
                height = int(img_height * ratio)
                width = int(img_width * ratio)
                new_img = pil_img.resize((width, height))
                new_img.fp = fp
            else:
                new_img = pil_img
            img = drawing.image.Image(new_img)
            img.anchor = photo_location[n]
            pages_list[i].add_image(img)
        ws.print_area = 'A1:D50'
    wb.remove(ws_del)

    file = BytesIO()
    wb.save(filename=file)
    wb.close()

    return file
