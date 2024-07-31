"""
ReportName: 屏東獨老季報
POC: Shen Chiang
"""

import json
from datetime import datetime
from functools import reduce
from io import BytesIO
from os import path

import pandas as pd
from bson import ObjectId
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from linkinpark.app.ds.reportPlatformBackend.utils import (ReportGenerateError,
                                                           check_org_type,
                                                           count_age,
                                                           get_nis_data,
                                                           trans_timezone,
                                                           trans_to_date)

REPLACE_SYMBOL = {"[": "", "]": "", "'": ""}
COUNTY = {
    "屏東市": 1, "潮州鎮": 2, "東港鎮": 3, "恆春鎮": 4, "萬丹鄉": 5,
    "長治鄉": 6, "麟洛鄉": 7, "九如鄉": 8, "里港鄉": 9, "鹽埔鄉": 10,
    "高樹鄉": 11, "萬巒鄉": 12, "內埔鄉": 13, "竹田鄉": 14, "新埤鄉": 15,
    "枋寮鄉": 16, "新園鄉": 17, "崁頂鄉": 18, "林邊鄉": 19, "南州鄉": 20,
    "佳冬鄉": 21, "琉球鄉": 22, "車城鄉": 23, "滿州鄉": 24, "枋山鄉": 25,
    "三地門鄉": 26, "霧台鄉": 27, "瑪家鄉": 28, "泰武鄉": 29, "來義鄉": 30,
    "春日鄉": 31, "獅子鄉": 32, "牡丹鄉": 33
}


def query_data_of_patients(
        org: ObjectId, start: datetime, end: datetime, suffix: dict = None):
    """
    This is a internal function for Ping-Tung home care community report. This
    function will query the patient data with care plan and service records.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: An external settings for the criteria of patients inclusion.
    :return: A dataframe of client and service details.
    """
    # Regulation of this report
    org_info = get_nis_data(
        "organizations", {"_id": org}, {"solution": 1, "addressCity": 1}
    ).to_dict(orient="index")[0]
    if org_info["solution"] != "homecare":
        raise ReportGenerateError(
            f"Not applicable for organization type {org_info['solution']}")
    if org_info["addressCity"] != "屏東縣":
        raise ReportGenerateError(
            f"Not applicable for organization addressCity "
            f"{org_info['addressCity']}")

    # parameter setting section
    if suffix:
        service_type = suffix["serviceType"]
        living_status = suffix["living_status"]
    else:
        service_type = ["homeCare"]
        living_status = ["alone"]
    all_living_options = [
        "alone",
        "withSpouse",
        "other",
        "family",
        "institution",
        "Government"
    ]
    exclude_options = list(set(all_living_options) - set(living_status))
    collections = [
        "patients",
        "punchclockrecords",
        "daycareservices",
        "approvedcareplans",
        "transfermanages",
        "telerecords",
        "homerecords",
    ]
    conditions = [
        {"organization": org,
         "isDeleted": {"$ne": True},
         "DCHCServiceType": {"$in": service_type},
         "$and": [
             {"livingStatus": {"$in": living_status}},
             {"livingStatus": {"$nin": exclude_options}}
         ]},
        {"organization": org,
         "date": {"$gte": start, "$lt": end}},
        {"organization": org},
        {"organization": org,
         "dateOfApproval": {'$lt': end},
         "planType": {"$ne": "termination"}},
        {"organization": org,
         "createdDate": {'$lt': end}},
        {"organization": org,
         "date": {"$gte": start, "$lt": end}},
        {"organization": org,
         "date": {"$gte": start, "$lt": end}}
    ]
    columns = [
        {"birthday": 1,
         "numbering": 1,
         "caseNumber": 1,
         "sex": 1,
         "residentialAddressArea": 1,
         "DCHCServiceType": 1,
         "lastName": 1,
         "firstName": 1,
         "handicapHandbook": 1,
         "livingStatus": 1},
        {"patient": 1, "date": 1, "services": 1, "serviceStatus": 1},
        {"code": 1},
        {"CMSLevel": 1,
         "dateOfApproval": 1,
         "socialWelfareStatus": 1,
         "socialWelfare": 1,
         "patient": 1,
         "disability": 1},
        {"patient": 1, "createdDate": 1, "status": 1, "firstServerDate": 1},
        {"_id": 1, "patient": 1, "date": 1},
        {"_id": 1, "patient": 1, "date": 1}
    ]

    # Querying section
    result = get_nis_data(collections, conditions, columns)

    # Preprocess Section
    for position in range(len(result)):
        result[position].name = collections[position]
        for col in columns[position]:
            if col not in result[position].columns:
                result[position][col] = None
    (patient, records, services, care_plan, status, tel, visit) = result

    patient = patient.rename(columns={"_id": "patient"})
    patient["age"] = count_age(patient["birthday"], start)

    records = records.explode("services").rename(
        columns={"_id": "record_id"}).reset_index(drop=True)
    if not records.empty:
        records = pd.merge(
            records, records["services"].apply(pd.Series),
            left_index=True, right_index=True)
    else:
        records[["service", "record_id"]] = [None, None]

    services = services.rename(columns={"_id": "service"})
    services["category"] = services["code"].str.slice(0, 2)
    services = services[services["category"] == "BA"]

    care_plan = care_plan.loc[
        care_plan.sort_values("dateOfApproval").groupby("patient")[
            "dateOfApproval"].idxmax()
    ].reset_index(drop=True)

    status_df = []
    for status_name, col_name in {
        "startServer": "start_at", "closed": "close_at"
    }.items():
        temp_df = status.loc[
            status["status"] == status_name, ["patient", "createdDate"]
        ].rename(columns={"createdDate": col_name})
        status_df.append(
            temp_df.loc[temp_df.groupby("patient")[col_name].idxmax()])
    temp_df = status.loc[
        status["status"] == "startServer",
        ["patient", "firstServerDate"]
    ].rename(columns={"firstServerDate": "serve_at"})
    status_df.append(
        temp_df.loc[temp_df.groupby("patient")["serve_at"].idxmax()])
    status_df.append(
        status.loc[
            status.groupby("patient")["createdDate"].idxmax(),
            ["patient", "status"]
        ].rename(columns={"status": "last_status"}))
    status_df = reduce(
        lambda l_df, r_df: pd.merge(l_df, r_df, on=["patient"], how="outer"),
        status_df)
    tel.rename(columns={"_id": "record_id"}, inplace=True)
    visit.rename(columns={"_id": "record_id"}, inplace=True)

    # Merging section
    records = pd.merge(records, services, on="service")
    patient = pd.merge(patient, care_plan, "left", on="patient")
    patient = pd.merge(patient, status_df, "left", on="patient")

    # Computing section
    patient = patient[
        (patient["start_at"] < end)
        & ((patient["close_at"] >= start) | pd.isna(patient["close_at"]))]

    index_all = records["record_id"].unique()
    index_ba14 = records[records["code"] == "BA14"]["record_id"].unique()
    index_not_ba14 = records[
        ~records["record_id"].isin(index_ba14)]["record_id"].unique()
    records_conditions = {
        # record type: [filter_by_serviceStatus, filter_by_service_code]
        "home_service": [["normal", "unmet"], index_not_ba14],
        "to_hospital": [["normal", "unmet"], index_ba14],
        "staff_away": [["staffAway"], index_all],
        "patient_away": [["patientAway"], index_all]
    }
    records_sub_df_dict = {
        "status_record": status, "tel_record": tel, "home_visit": visit}
    for record_name, conditions_filter in records_conditions.items():
        service_status, index = conditions_filter
        records_sub_df_dict[record_name] = records[
            (records["serviceStatus"].isin(service_status))
            & (records["record_id"].isin(index))]

    result = [patient]
    for record_name, df in records_sub_df_dict.items():
        show_amount = True
        try:
            if df.name == "transfermanages":
                df["date"], df["record_id"] = df["createdDate"], df["status"]
                show_amount = False
        except AttributeError:
            pass
        df = df.sort_values("date")
        df["display_date"] = (
            pd.to_datetime(df["date"]) + pd.Timedelta(hours=8)).dt.date
        if show_amount:
            df = df.groupby(["patient", "display_date"]).agg(
                amount=("record_id", "nunique")).reset_index()
        else:
            df = df.groupby(["patient", "display_date", "record_id"]).agg(
                amount=("record_id", (lambda x: x.unique()))).reset_index()
        df["display"] = (
            pd.to_datetime(df["display_date"]).astype(str)
            + " (" + df["amount"].astype(str) + ")")
        temp_result = {}
        for _id in patient["patient"]:
            temp_df = df[df["patient"] == _id]
            temp_result[_id] = [
                temp_df["amount"].sum(), ",\n".join(temp_df["display"])]
        result.append(
            pd.DataFrame.from_dict(
                temp_result, orient="index"
            ).reset_index().rename(
                columns={"index": "patient",
                         0: record_name + "_amount",
                         1: record_name + "_content"}))
    result = reduce(
        lambda l_df, r_df: pd.merge(l_df, r_df, on=["patient"], how="left"),
        result)
    result.loc[
        (result["serve_at"] >= start) & (result["serve_at"] < end),
        "new_comer"
    ] = 1
    return result


def generate_data_of_sheet_1(df):
    # parameter setting section
    service_coordinate = {
        "home_visit_amount": 6,
        "tel_record_amount": 9,
        "to_hospital_amount": 12,
        "home_service_amount": 15,
        "new_comer": 18
    }
    sex_coordinate = {
        "male": 1,
        "female": 2,
    }
    axis_y_coordinate = {}
    for county, position in COUNTY.items():
        axis_y_coordinate[county] = position * 5 + 8

    # Computing section
    result = {}
    for county, row in axis_y_coordinate.items():
        for amount_type, service_pos in service_coordinate.items():
            for sex, sex_pos in sex_coordinate.items():
                col = get_column_letter(service_pos + sex_pos)
                result[col + str(row)] = df.loc[
                    (df["residentialAddressArea"] == county)
                    & (df["sex"] == sex),
                    amount_type].sum()
    return result


def generate_data_of_sheet_2(df):
    # parameter setting section
    replace_dict = {
        "sex": {"female": "女", "male": "男"},
        "last_status": {
            "startServer": "服務中",
            "continue": "服務中",
            "bedTransfer": "服務中",
            "branchTransfer": "服務中",
            "closed": "結案",
            "pause": "暫停服務"},
        "DCHCServiceType": {
            **REPLACE_SYMBOL,
            "HCOwnExpense": "自費",
            "HCRespite": "居家喘息",
            "homeCare": "居家服務",
            "HCShortcare": "居家短照",
        },
        "livingStatus": {
            **REPLACE_SYMBOL,
            "alone": "獨居",
            "family": "與家人同住",
            "withMate": "與配偶住",
            "institution": "住宿式機構",
            "Government": "政府補助居住服務",
            "other": "與他人同住",
        },
        "status_record_content": {
            "(startServer)": "開始服務",
            "(branchTransfer)": "換區域/組別",
            "(pause)": "暫停",
            "(continue)": "繼續服務",
            "(closed)": "結案",
            "(closedTracking)": "結案追蹤",
        }
    }

    # Preprocess Section
    df["display_name"] = (
        "[" + df["residentialAddressArea"].str.replace("區", "") + "]"
        + df["numbering"].fillna("") + " "
        + df["lastName"] + df["firstName"])
    for col, replace in replace_dict.items():
        df[col] = df[col].astype(str)
        for pat, repl in replace.items():
            df[col] = df[col].str.replace(pat, repl, regex=False)
    df["birthday"] = df["birthday"].dt.strftime("%Y-%m-%d")
    for col in ("start_at", "close_at", "serve_at"):
        df[col] = (df[col] + pd.Timedelta(hours=8)).dt.strftime("%Y-%m-%d")
    columns = {
        "display_name": "個案",
        "sex": "性別",
        "age": "年齡",
        "birthday": "生日",
        "last_status": "服務狀態",
        "DCHCServiceType": "服務類別",
        "residentialAddressArea": "居住區域",
        "livingStatus": "居住狀況",
        "start_at": "開案日期",
        "serve_at": "開始服務日期",
        "close_at": "結案日期",
        "status_record_content": "異動紀錄",
        "tel_record_amount": "電話問安次數",
        "tel_record_content": "電訪日期",
        "home_visit_amount": "關懷訪視次數",
        "home_visit_content": "家訪日期",
        "home_service_amount": "居家服務次數",
        "home_service_content": "居家服務日期",
        "to_hospital_amount": "陪同就醫次數",
        "to_hospital_content": "陪同就醫日期",
        "staff_away_amount": "人員請假次數",
        "staff_away_content": "人員請假日期",
        "patient_away_amount": "個案請假次數",
        "patient_away_content": "個案請假日期",
    }
    result = df[columns.keys()].rename(columns=columns)
    return result


def generate_data_of_sheet_3(df1, df2, time_list):
    # parameter setting section
    now, _, past, _ = time_list
    quarter_now, quarter_past = [
        f"{x.year}Q{(x.month - 1) // 3 + 1}" for x in (now, past)
    ]
    amount_coordinate = {
        "total_amount": 2,
        "people": 4,
        "staff_away_amount": 6,
        "patient_away_amount": 8,
    }
    quarter_coordinate = {quarter_past: 0, quarter_now: 1}
    axis_y_coordinate = {}
    for county, position in COUNTY.items():
        axis_y_coordinate[county] = position + 3

    # Preprocess Section
    df = pd.DataFrame()
    for data, quarter in zip((df1, df2), (quarter_now, quarter_past)):
        data["quarter"] = quarter
        df = pd.concat([df, data])
    df["total_amount"] = df[[
        "tel_record_amount",
        "home_visit_amount",
        "home_service_amount",
        "to_hospital_amount"
    ]].sum(axis=1)
    df["people"] = 1

    # Computing section
    result = {}
    for amount_type, col_main in amount_coordinate.items():
        for quarter, col_add in quarter_coordinate.items():
            col = get_column_letter(col_main + col_add)
            for county, row in axis_y_coordinate.items():
                result[col + str(row)] = df[
                    (df["quarter"] == quarter)
                    & (df["residentialAddressArea"] == county)][
                    amount_type].sum()
    for cols, quarter in {("B", "D", "F", "H"): quarter_past,
                          ("C", "E", "G", "I"): quarter_now}.items():
        for col in cols:
            result[col + "3"] = quarter
    return result


def pingtung_solitary_elder_report(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a statistic report of home care service
    provided by institution at Ping-Tung.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    if not check_org_type(org, "homecare"):
        raise ReportGenerateError("此為居服專屬報表，無法應用於其他機構類型。")

    # parameter setting section
    if isinstance(suffix, str):
        suffix_dict = json.loads(suffix)
    else:
        suffix_dict = None

    input_dates = trans_to_date([start, end])
    all_dates = [*input_dates]
    for input_date in input_dates:
        all_dates.append(input_date - relativedelta(months=3))
    start_now, end_now, start_past, end_past = trans_timezone(
        all_dates, from_utc=8, to_utc=0,
    )
    now = (org, start_now, end_now, suffix_dict)
    past = (org, start_past, end_past, suffix_dict)
    df_now = query_data_of_patients(*now)
    df_past = query_data_of_patients(*past)
    sheet1_result = generate_data_of_sheet_1(df_now)
    sheet2_result = generate_data_of_sheet_2(df_now)
    sheet3_result = generate_data_of_sheet_3(df_now, df_past, all_dates)

    # Filing section
    report_file = BytesIO()
    f_path = path.dirname(path.dirname(path.abspath(__file__)))
    try:
        work_book = load_workbook(
            "pipelines/ds_reportplatform_generator/templates/"
            "pch_hc_solitary_elder_template.xlsx"
        )
    except FileNotFoundError:
        work_book = load_workbook(
            path.join(
                f_path, "templates", "pch_hc_solitary_elder_template.xlsx"))
    for sheet_name, value_dict in {
        "長照處": sheet1_result,
        "季變化比較": sheet3_result,
    }.items():
        work_sheet = work_book[sheet_name]
        for loc, value in value_dict.items():
            work_sheet[loc] = value

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = work_book
        writer.sheets = dict((ws.title, ws) for ws in work_book.worksheets)

        # Fill in data to sheet 4
        sheet2_result.to_excel(
            writer,
            sheet_name="個案清單",
            header=False,
            index=False,
            startrow=1)

    return report_file
