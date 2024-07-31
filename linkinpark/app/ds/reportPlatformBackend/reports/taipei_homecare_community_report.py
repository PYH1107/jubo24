"""
ReportName: 台北居服社區月報
POC: Shen Chiang
"""

import json
from datetime import datetime
from functools import reduce
from io import BytesIO
from os import path

import pandas as pd
from bson.objectid import ObjectId
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from linkinpark.app.ds.reportPlatformBackend.utils import (ReportGenerateError,
                                                           check_org_type,
                                                           count_age,
                                                           get_nis_data,
                                                           preprocess_date)

REPLACE_SYMBOL = {"[": "", "]": "", "'": ""}
USER_COLUMNS = {
    "displayName": 1,
    "employeeNumber": 1,
    "jobTitle": 1,
    "jobType": 1,
    "employType": 1,
    "staffStatus": 1,
    "sex": 1,
    "resignDate": 1,
    "competencyCertification": 1
}


def query_data_of_patients(
        org: ObjectId, start: datetime, end: datetime, suffix=None):
    """
    This is a internal function for taipei_homecare_community_report. This
    function will query the patient data with care plan and service records,
    which will be used in sheet 1, 2, 3 and 4 in the Taipei home-care
    community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: An external settings for whether respite service only
    clients should be included or not.
    :return: A dataframe of client and service details.
    """
    # parameter setting section
    if suffix:
        service_type = suffix["serviceType"]
    else:
        service_type = ["homeCare", "HCRespite"]
    collections = [
        "patients",
        "punchclockrecords",
        "daycareservices",
        "approvedcareplans",
        "transfermanages",
    ]
    conditions = [
        {"organization": org,
         "DCHCServiceType": {"$in": service_type}
         },
        {"organization": org,
         "date": {"$gte": start, "$lt": end},
         "serviceStatus": {"$in": ["normal", "unmet"]}},
        {"organization": org},
        {"organization": org,
         "dateOfApproval": {'$lt': end},
         "planType": {"$ne": "termination"}},
        {"organization": org,
         "createdDate": {'$lt': end}},
    ]
    columns = [
        {"birthday": 1,
         "numbering": 1,
         "caseNumber": 1,
         "sex": 1,
         "residentialAddressArea": 1,
         "DCHCServiceType": 1,
         "lastName": 1,
         "firstName": 1,
         "handicapHandbook": 1,
         "livingStatus": 1},
        {"patient": 1, "date": 1, "services": 1},
        {"code": 1},
        {"CMSLevel": 1,
         "dateOfApproval": 1,
         "socialWelfareStatus": 1,
         "socialWelfare": 1,
         "patient": 1,
         "disability": 1},
        {"patient": 1, "createdDate": 1, "status": 1},
    ]

    # Querying section
    patient, records, services, care_plan, status = get_nis_data(
        collections, conditions, columns
    )

    # Preprocess Section
    patient = patient.rename(columns={"_id": "patient"})
    for col in columns[0]:
        if col not in patient.columns:
            patient[col] = None
    patient["age"] = count_age(patient["birthday"], start)
    records = records.explode("services").rename(
        columns={"_id": "record_id"}).reset_index(drop=True)
    records = pd.merge(
        records,
        records["services"].apply(pd.Series),
        left_index=True,
        right_index=True
    )

    services = services.rename(columns={"_id": "service"})
    services["category"] = services["code"].str.slice(0, 2)
    services = services[services["category"].isin(["BA", "GA"])]
    care_plan = care_plan.loc[
        care_plan.sort_values("dateOfApproval").groupby("patient")[
            "dateOfApproval"].idxmax()
    ].reset_index(drop=True)

    start_df = status.loc[
        status["status"] == "startServer",
        ["patient", "createdDate"]
    ].rename(columns={"createdDate": "start_at"})
    start_df = start_df.loc[
        start_df.groupby("patient")["start_at"].idxmax()
    ]

    close_df = status.loc[
        status["status"] == "closed",
        ["patient", "createdDate"]
    ].rename(columns={"createdDate": "close_at"})
    close_df = close_df.loc[
        close_df.groupby("patient")["close_at"].idxmax()
    ]

    last_df = status.loc[status.groupby("patient")["createdDate"].idxmax()]
    last_df = last_df[
        ["patient", "status"]
    ].rename(columns={"status": "last_status"})
    status = reduce(
        lambda left, right: pd.merge(
            left, right, on=['patient'], how='outer'),
        [start_df, close_df, last_df]
    )

    # Merging section
    records = pd.merge(records, services, on="service")
    patient = pd.merge(patient, care_plan, "left", on="patient")
    patient = pd.merge(patient, status, "left", on="patient")

    # Computing section
    patient = patient[
        (patient["start_at"] < end)
        & ((patient["close_at"] >= start)
           | pd.isna(patient["close_at"]))]
    amount_df = records.groupby("patient").agg(
        {"record_id": "nunique"}).reset_index().rename(
        columns={"record_id": "amount"})
    records = records.groupby(["patient", "category", "date"]).agg(
        {"record_id": "nunique"}).reset_index()
    records["present"] = (
        records["date"].dt.date.astype(str)
        + "(" + records["record_id"].astype(str) + ")")
    records_dict = {}
    for _id in records["patient"].unique():
        records_dict[_id] = {}
        for cat in ("BA", "GA"):
            date = list(
                records.loc[
                    (records["patient"] == _id)
                    & (records["category"] == cat),
                    "present"])
            records_dict[_id][cat] = sorted(date)
    records = pd.DataFrame.from_dict(
        records_dict, orient="index"
    ).reset_index().rename(columns={"index": "patient"})
    result = pd.merge(patient, records, "left", on="patient")
    result = pd.merge(result, amount_df, "left", on="patient")
    return result


def generate_data_of_sheet_1(df, start: datetime):
    """
    This is a internal function for taipei_homecare_community_report. This
    function is used to generate the data result for the 1st sheet in Taipei
    home-care community report.
    :param df: The dataframe of patients.
    :param start: The start date of report period.
    :return: A dictionary of the position and value.
    """
    # parameter setting section
    gender_coordinate = {"male": 1, "female": 2}
    status_coordinate = {
        "新案": 5, "舊案": 8, "結案": 11, "當月開結": 14, "暫停": 17
    }
    social_coordinate = {
        "lowInLaw": "D",
        "lowToMid": "E",
        "low": "D",
        "normal": "F",
        "nan": "G",
    }

    # Preprocess Section
    df["socialWelfare"] = df["socialWelfare"].astype(str)
    for pat, repl in REPLACE_SYMBOL.items():
        df["socialWelfare"] = df["socialWelfare"].str.replace(
            pat, repl, regex=False
        )
    df.loc[df["start_at"] >= start, "新案"] = "新案"
    df.loc[
        (df["start_at"] < start) & (df["last_status"] != "closed"),
        "舊案"
    ] = "舊案"
    df.loc[df["close_at"] >= start, "結案"] = "結案"
    df.loc[
        (df["close_at"] >= start) & (df["start_at"] >= start),
        "當月開結"
    ] = "當月開結"
    df.loc[df["last_status"] == "pause", "暫停"] = "暫停"
    df["status"] = ""
    for status in status_coordinate:
        df["status"] += (df[status].fillna("") + ",")
    df["status"] = df["status"].str.split(",")

    # Computing section
    result = {}
    for social, col in social_coordinate.items():
        for status, status_row in status_coordinate.items():
            for sex, gender_row in gender_coordinate.items():
                result[col + str(status_row + gender_row)] = len(
                    df.loc[
                        (df["socialWelfare"] == social)
                        & (df["status"].map({status}.issubset))
                        & (df["sex"] == sex)])
    for social, col in social_coordinate.items():
        for sex, gender_row in gender_coordinate.items():
            result[col + str(gender_row + 20)] = df.loc[
                (df["socialWelfare"] == social) & (df["sex"] == sex),
                "amount"].sum()
    return result


def generate_data_of_sheet_2(df):
    """
    This is a internal function for taipei_homecare_community_report. This
    function is used to generate the data result for the 2nd sheet in Taipei
    home-care community report.
    :param df: The dataframe of patients.
    :return: A dictionary of the position and value.
    """
    # parameter setting section
    gender_coordinate = {"male": 1, "female": 2}
    area_coordinate = {
        "松山區": 5,
        "信義區": 8,
        "大安區": 11,
        "中山區": 14,
        "中正區": 17,
        "大同區": 20,
        "萬華區": 23,
        "文山區": 26,
        "南港區": 29,
        "內湖區": 32,
        "士林區": 35,
        "北投區": 38,
        "無法分類": 41,
    }
    age_group = {"50-64": "D", "65-74": "E", "75-84": "F", "85以上": "G"}

    # Preprocess Section
    df.loc[
        ~df["residentialAddressArea"].isin(area_coordinate),
        "residentialAddressArea"] = "無法分類"
    df["age_group"] = pd.cut(
        df["age"],
        [50, 65, 75, 85, float("inf")],
        labels=["50-64", "65-74", "75-84", "85以上"],
        right=False)

    # Computing section
    result = {}
    for age, col in age_group.items():
        for area, area_row in area_coordinate.items():
            for sex, gender_row in gender_coordinate.items():
                result[col + str(area_row + gender_row)] = len(
                    df[(df["age_group"] == age)
                       & (df["residentialAddressArea"] == area)
                       & (df["sex"] == sex)])
    return result


def generate_data_of_sheet_3(df):
    """
    This is a internal function for taipei_homecare_community_report. This
    function is used to generate the data result for the 3rd sheet in Taipei
    home-care community report.
    :param df: The dataframe of patients.
    :return: A dictionary of the position and value.
    """
    # parameter setting section
    gender_coordinate = {"male": 1, "female": 2}
    living_coordinate = {
        "alone": 6,
        "withMate": 9,
        "family": 12,
        "other": 15,
        "other_option": 18,
        "unknown": 21}
    age_coordinate = {"below64": 7, "above65": 6}
    disability_coordinate = {"yes": 0, "no": -1}

    # Preprocess Section
    df = df[df["age"] > 49]
    df["age_cat"] = "below64"
    df.loc[df["age"] >= 65, "age_cat"] = "above65"
    df["live_cat"] = "other_option"
    for status in living_coordinate.keys():
        df.loc[
            df["livingStatus"].astype(str) == f"['{status}']",
            "live_cat"] = status
    df.loc[df["livingStatus"].apply(len) < 1, "live_cat"] = "unknown"
    df.loc[df["handicapHandbook"] != "yes", "handicapHandbook"] = "no"

    # Computing section
    result = {}
    for age_cat, age_col in age_coordinate.items():
        for hand_book, book_col in disability_coordinate.items():
            for living, live_row in living_coordinate.items():
                for sex, sex_row in gender_coordinate.items():
                    col = get_column_letter(age_col + book_col)
                    row = live_row + sex_row
                    result[col + str(row)] = len(
                        df[(df["age_cat"] == age_cat)
                           & (df["handicapHandbook"] == hand_book)
                           & (df["live_cat"] == living)
                           & (df["sex"] == sex)])
    return result


def generate_data_of_sheet_4(df):
    """
    This is a internal function for taipei_homecare_community_report. This
    function is used to generate the data result for the 4th sheet in Taipei
    home-care community report.
    :param df: The dataframe of patients.
    :return: A dataframe to be filled in sheet 4.
    """
    # parameter setting section
    replace_dict = {
        "sex": {"female": "女", "male": "男"},
        "last_status": {
            "startServer": "服務中",
            "continue": "服務中",
            "bedTransfer": "服務中",
            "branchTransfer": "服務中",
            "closed": "結案",
            "pause": "暫停服務"},
        "DCHCServiceType": {
            **REPLACE_SYMBOL,
            "HCOwnExpense": "自費",
            "HCRespite": "居家喘息",
            "homeCare": "居家服務",
        },
        "socialWelfare": {
            **REPLACE_SYMBOL,
            "lowInLaw": "長照低收",
            "lowToMid": "長照中低收",
            "low": "長照低收",
            "normal": "一般戶",
            "nan": "未填寫",
        },
        "livingStatus": {
            **REPLACE_SYMBOL,
            "alone": "獨居",
            "family": "與家人同住",
            "withMate": "與配偶住",
            "institution": "住宿式機構",
            "Government": "政府補助居住服務",
            "other": "與他人同住",
        },
        "handicapHandbook": {
            "yes": "是", "no": "否", "applying": "申請中", "nan": "未填寫"
        },
        "disability": {"yes": "是", "no": "否", "nan": "未填寫"},
        "BA": {**REPLACE_SYMBOL, "nan": ""},
        "GA": {**REPLACE_SYMBOL, "nan": ""},
    }

    # Preprocess Section
    df["display_name"] = (
        "[" + df["residentialAddressArea"].str.replace("區", "") + "]"
        + df["numbering"].fillna("") + " "
        + df["lastName"] + df["firstName"])
    for col, replace in replace_dict.items():
        df[col] = df[col].astype(str)
        for pat, repl in replace.items():
            df[col] = df[col].str.replace(pat, repl, regex=False)
    df["birthday"] = df["birthday"].dt.strftime("%Y-%m-%d")
    for col in ("start_at", "close_at"):
        df[col] = (df[col] + pd.Timedelta(hours=8)).dt.strftime("%Y-%m-%d")
    columns = {
        "display_name": "個案",
        "sex": "性別",
        "age": "年齡",
        "birthday": "生日",
        "last_status": "服務狀態",
        "DCHCServiceType": "服務類別",
        "residentialAddressArea": "居住區域",
        "livingStatus": "居住狀況",
        "handicapHandbook": "身心障礙證明/手冊",
        "CMSLevel": "CMS等級",
        "socialWelfare": "長照福利身份",
        "disability": "身心障礙-失智",
        "start_at": "開案日期",
        "close_at": "結案日期",
        "BA": "BA碼打卡日期",
        "GA": "GA碼打卡日期",
    }
    result = df[columns.keys()].rename(columns=columns)
    return result


def generate_data_of_sheet_5(org: ObjectId, start: datetime):
    """
    This is a internal function for taipei_homecare_community_report. This
    function is used to generate the data result for the 5th sheet in Taipei
    home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :return: A dictionary of the position and value.
    """
    # parameter setting section
    gender_coordinate = {"male": "C", "female": "D"}
    job_coordinate = {
        "3": ("supervisor", ["fullTime", "partTime"]),
        "6": ("homeAttendant", ["fullTime"]),
        "7": ("homeAttendant", ["partTime"])}
    train_coordinate = {
        "SpecialEducationTraining": "9",
        "footCareCourse": "10",
        "dementiaClass": "11",
        "disabilityCertificate": "12"}

    # Querying section
    df = get_nis_data(
        "users",
        {"organization": org,
         "jobType": {"$in": ["homeAttendant", "supervisor"]},
         "staffStatus": {"$ne": "resigned"},
         "$or": [{"resignDate": None}, {"resignDate": {"$gte": start}}]},
        USER_COLUMNS)

    # Computing section
    result = {}
    for gender, col in gender_coordinate.items():
        for row, job in job_coordinate.items():
            job_type, employ_type = job
            result[col + row] = len(
                df[(df["sex"] == gender)
                   & (df["jobType"] == job_type)
                   & (df["employType"].isin(employ_type))])
        for certificate, row in train_coordinate.items():
            result[col + row] = len(
                df[
                    (df["sex"] == gender)
                    & (df["competencyCertification"].str.contains(
                        certificate, regex=False))])
    return result


def generate_data_of_sheet_6(org: ObjectId, start: datetime):
    """
    This is a internal function for taipei_homecare_community_report. This
    function is used to generate the data result for the 6th sheet in Taipei
    home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :return: A dataframe to be filled in sheet 6.
    """
    # Querying section
    df = get_nis_data(
        "users",
        {"organization": org,
         "jobType": {"$in": ["homeAttendant", "supervisor"]},
         "$or": [{"resignDate": None}, {"resignDate": {"$gte": start}}]},
        USER_COLUMNS,
    )
    # Preprocess Section
    errors = []
    for col in USER_COLUMNS:
        if col not in df.columns:
            df[col] = None
    df["displayName"] = (
        df["employeeNumber"].fillna("") + " " + df["displayName"].fillna(""))
    df["displayStatus"] = df["staffStatus"]
    for needed_col in ["jobType", "employType", "sex"]:
        if any(df[needed_col].isna()):
            name_list = df[df[needed_col].isna()]['displayName'].to_list()
            errors.append(f"{needed_col} missing: {', '.join(name_list)}")
    if len(errors) > 0:
        for error in errors:
            print(error, flush=True)
        raise ReportGenerateError(
            "<br>".join(["人員檔案內缺少必要的欄位資訊：", *errors]))
    df["competencyCertification"] = df["competencyCertification"].astype(str)
    result = df.replace({
        "sex": {"male": "男", "female": "女"},
        "jobType": {"homeAttendant": "居家服務員", "supervisor": "督導員"},
        "employType": {"fullTime": "全職", "partTime": "兼職"},
        "displayStatus": {"employed": "在職",
                          "furlough": "留職停薪",
                          "resigned": "離職"}})
    result["resignDate"] = pd.to_datetime(result["resignDate"])
    result.loc[
        result["staffStatus"] == "resigned", "displayStatus"
    ] += result["resignDate"].dt.strftime("(%Y-%m-%d)")
    replace_dict = {
        **REPLACE_SYMBOL,
        "dementiaClass": "失智症照顧服務20小時訓練課程",
        "disabilityCertificate": "身心障礙支持服務核心課程訓練",
        "attendantCertificate": "照顧服務員單一級技術士技能證照",
        "footCareCourse": "足部照護課程",
        "SpecialEducationTraining": "口腔內及人工氣道管內分泌物之清潔、抽吸與移除"}
    for k, v in replace_dict.items():
        result["competencyCertification"] = result[
            "competencyCertification"].str.replace(k, v, regex=True)
    result = result[[
        "displayName",
        "sex",
        "jobType",
        "jobTitle",
        "employType",
        "displayStatus",
        "competencyCertification"]]
    return result


def generate_data_of_sheet_7(org: ObjectId, start: datetime, end: datetime):
    """
    This is a internal function for taipei_homecare_community_report. This
    function is used to generate the data result for the 7th sheet in Taipei
    home-care community report.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :return: A dictionary of the position and value.
    """
    # parameter setting section
    coordinate = {
        "AA03": "B2", "AA04": "B3", "AA05": "B4",
        "AA06": "B5", "AA07": "B6", "AA08": "B7",
        "AA09": "B8", "AA10": "B9", "AA11": "B10",
    }

    # Querying section
    df = get_nis_data(
        "importedservices",
        {"organization": org,
         "date": {"$gte": start, "$lt": end}},
        {"code": 1})

    # Computing section
    if not df.empty:
        df = df.groupby("code").count()
    result = {}
    for service_code, position in coordinate.items():
        if service_code in df.index:
            result[position] = df.at[service_code, "_id"]
        else:
            result[position] = 0
    return result


def taipei_homecare_community_report(
        org: ObjectId, start: datetime, end: datetime, suffix: str = None
) -> BytesIO:
    """
    This function will generate a statistic report of home care service
    provided by institution at Taipei.
    :param org: The id of organization for this report.
    :param start: Start date of report period.
    :param end: End date of report period.
    :param suffix: Not in use for this function.
    :return: The report generated.
    """

    if not check_org_type(org, "homecare"):
        raise ReportGenerateError("此為居服專屬報表，無法應用於其他機構類型。")

    # parameter setting section
    if isinstance(suffix, str):
        suffix_dict = json.loads(suffix)
    else:
        suffix_dict = None
    start, end, query_start, query_end = preprocess_date([start, end])

    # Querying section
    df = query_data_of_patients(org, query_start, query_end, suffix_dict)
    sheet1_result = generate_data_of_sheet_1(df, query_start)
    sheet2_result = generate_data_of_sheet_2(df)
    sheet3_result = generate_data_of_sheet_3(df)
    sheet4_result = generate_data_of_sheet_4(df)
    sheet5_result = generate_data_of_sheet_5(org, query_start)
    sheet6_result = generate_data_of_sheet_6(org, query_start)
    sheet7_result = generate_data_of_sheet_7(org, query_start, query_end)

    # Filing section
    report_file = BytesIO()
    f_path = path.dirname(path.dirname(path.abspath(__file__)))
    try:
        work_book = load_workbook(
            "pipelines/ds_reportplatform_generator/templates/"
            "tpe_hc_community_template.xlsx")
    except FileNotFoundError:
        work_book = load_workbook(
            path.join(f_path, "templates", "tpe_hc_community_template.xlsx"))
    for sheet_name, value_dict in {
        "個案-服務人數人次": sheet1_result,
        "個案-各區域在案人數": sheet2_result,
        "個案-各居住狀況在案人數": sheet3_result,
        "人員": sheet5_result,
        "AA碼數量": sheet7_result
    }.items():
        work_sheet = work_book[sheet_name]
        for loc, value in value_dict.items():
            work_sheet[loc] = value

    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        writer.book = work_book
        writer.sheets = dict((ws.title, ws) for ws in work_book.worksheets)

        # Fill in data to sheet 4
        sheet4_result.to_excel(
            writer,
            sheet_name="個案清單",
            header=False,
            index=False,
            startrow=1)
        # Fill in data to sheet 6
        sheet6_result.to_excel(
            writer,
            sheet_name="人員清單",
            header=False,
            index=False,
            startrow=1)

    return report_file
